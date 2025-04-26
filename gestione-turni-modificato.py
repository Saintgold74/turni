import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime, timedelta
import calendar
# import random # Non più usato direttamente nella nuova logica, ma potrebbe servire altrove
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import traceback # Import aggiunto per debug dettagliato

class GestioneTurni:
    def __init__(self):
        """Inizializzazione dell'applicazione"""
        # Inizializzazione delle variabili principali
        self.addetti = {}  # Dizionario per memorizzare i dati degli addetti
        self.turni_disponibili = []  # Lista dei turni disponibili
        # Lista festività - potrebbe essere resa configurabile
        self.giorni_festivi = [
            "01-01",  # Capodanno
            "06-01",  # Epifania
            # Pasqua e Pasquetta richiedono calcolo
            "25-04",  # Liberazione
            "01-05",  # Festa dei Lavoratori
            "02-06",  # Festa della Repubblica
            "15-08",  # Ferragosto
            "01-11",  # Ognissanti
            "08-12",  # Immacolata Concezione
            "25-12",  # Natale
            "26-12"   # Santo Stefano
            # Aggiungere eventuali patroni locali o altre festività
        ]

        # Orari di apertura del supermercato - potrebbero essere resi configurabili
        self.orario_apertura = "08:00"
        self.orario_chiusura = "21:00"

        # Colori per Excel
        self.colori = {
            'header': 'CCE5FF',     # Azzurro chiaro per header
            'weekend': 'FFE6E6',    # Rosa chiaro per weekend
            'turno_mattina': 'E6FFE6',  # Verde chiaro per turni mattina
            'turno_pomeriggio': 'FFE6CC',  # Arancione chiaro per turni pomeriggio
            'riposo': 'F2F2F2',     # Grigio chiaro per riposi
            'ferie': 'FFFF99',      # Giallo chiaro per ferie
            'festivo': 'FF9999',    # Rosso chiaro per festivi
            'errore': 'FF0000'      # Rosso per errori/copertura incompleta
        }

        # Carica i dati se esistono
        self.carica_dati()

        # Creazione della finestra principale
        self.root = tk.Tk()
        self.root.title("Gestione Turni Supermercato")
        self.root.geometry("800x600") # Dimensione iniziale

        # Creazione del menu principale
        self.crea_menu_principale()

    def carica_dati(self):
        """Carica i dati salvati se esistono"""
        try:
            if os.path.exists('dati_turni.json'):
                with open('dati_turni.json', 'r', encoding='utf-8') as f:
                    dati = json.load(f)
                    self.addetti = dati.get('addetti', {})
                    # Assicura che ferie e riposi siano liste
                    for nome, info in self.addetti.items():
                         info['ferie'] = info.get('ferie', [])
                         info['giorni_riposo'] = info.get('giorni_riposo', [])
                    self.turni_disponibili = dati.get('turni', [])
            print("Dati caricati con successo.")
        except Exception as e:
            print(f"Errore nel caricamento dei dati: {e}")
            messagebox.showerror("Errore Caricamento", f"Impossibile caricare i dati da dati_turni.json.\n{e}")

    def salva_dati(self):
        """Salva i dati su file"""
        try:
            dati = {
                'addetti': self.addetti,
                'turni': self.turni_disponibili
            }
            with open('dati_turni.json', 'w', encoding='utf-8') as f:
                json.dump(dati, f, indent=4) # indent=4 per leggibilità
            print("Dati salvati con successo.")
        except Exception as e:
            print(f"Errore nel salvataggio dei dati: {e}")
            messagebox.showerror("Errore Salvataggio", f"Impossibile salvare i dati su dati_turni.json.\n{e}")

    def crea_menu_principale(self):
        """Crea il menu principale dell'applicazione"""
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(expand=True, fill=tk.BOTH) # Usa pack per centrare meglio

        # Stile per i bottoni
        style = ttk.Style()
        style.configure('TButton', font=('Helvetica', 12), padding=10)

        ttk.Button(main_frame, text="Gestione Addetti",
                   command=self.gestione_addetti, style='TButton').pack(pady=10, fill=tk.X)
        ttk.Button(main_frame, text="Gestione Turni",
                   command=self.gestione_turni, style='TButton').pack(pady=10, fill=tk.X)
        ttk.Button(main_frame, text="Gestione Ferie e Riposi",
                   command=self.gestione_ferie_riposi, style='TButton').pack(pady=10, fill=tk.X)
        ttk.Button(main_frame, text="Genera Pianificazione",
                   command=self.genera_pianificazione, style='TButton').pack(pady=10, fill=tk.X)
        ttk.Button(main_frame, text="Visualizza Statistiche",
                   command=self.visualizza_statistiche, style='TButton').pack(pady=10, fill=tk.X)

    def gestione_addetti(self):
        """Gestisce l'aggiunta e la modifica degli addetti"""
        # (Codice della funzione gestione_addetti originale)
        # ... (incolla qui il codice della funzione gestione_addetti dal file originale) ...
        window = tk.Toplevel(self.root)
        window.title("Gestione Addetti")
        window.geometry("600x500")

        # Frame per la lista degli addetti esistenti
        frame_lista = ttk.LabelFrame(window, text="Addetti", padding=10)
        frame_lista.grid(row=0, column=0, padx=10, pady=10, sticky='ns')

        # Lista degli addetti esistenti
        lista_addetti = tk.Listbox(frame_lista, width=30, height=15)
        lista_addetti.pack(side=tk.LEFT, fill=tk.Y)
        scrollbar = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=lista_addetti.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        lista_addetti.config(yscrollcommand=scrollbar.set)

        for addetto in sorted(self.addetti.keys()): # Ordina alfabeticamente
            lista_addetti.insert(tk.END, addetto)

        # Frame per il form di inserimento/modifica
        frame_form = ttk.LabelFrame(window, text="Dettagli Addetto", padding=10)
        frame_form.grid(row=0, column=1, padx=10, pady=10, sticky='nsew')

        # Configura grid per espansione
        window.grid_columnconfigure(1, weight=1)
        window.grid_rowconfigure(0, weight=1)

        ttk.Label(frame_form, text="Nome:").grid(row=0, column=0, pady=5, sticky='w')
        nome_var = tk.StringVar()
        nome_entry = ttk.Entry(frame_form, textvariable=nome_var, width=30)
        nome_entry.grid(row=0, column=1, pady=5, sticky='ew')

        ttk.Label(frame_form, text="Ore Contratto:").grid(row=1, column=0, pady=5, sticky='w')
        ore_var = tk.IntVar(value=40)
        ore_spinbox = ttk.Spinbox(frame_form, from_=0, to=100, textvariable=ore_var, width=5)
        ore_spinbox.grid(row=1, column=1, pady=5, sticky='w')

        ttk.Label(frame_form, text="Ore Max Settimanali:").grid(row=2, column=0, pady=5, sticky='w')
        ore_max_var = tk.IntVar(value=48)
        ore_max_spinbox = ttk.Spinbox(frame_form, from_=0, to=100, textvariable=ore_max_var, width=5)
        ore_max_spinbox.grid(row=2, column=1, pady=5, sticky='w')

        ttk.Label(frame_form, text="Straordinario Autorizzato:").grid(row=3, column=0, pady=5, sticky='w')
        straordinario_var = tk.BooleanVar()
        straordinario_check = ttk.Checkbutton(frame_form, variable=straordinario_var, onvalue=True, offvalue=False)
        straordinario_check.grid(row=3, column=1, pady=5, sticky='w')

        # Funzione per caricare dati dell'addetto selezionato
        def carica_dati_selezionato(event=None):
            selection = lista_addetti.curselection()
            if selection:
                nome = lista_addetti.get(selection[0])
                if nome in self.addetti:
                    info = self.addetti[nome]
                    nome_var.set(nome)
                    ore_var.set(info.get('ore_contratto', 40))
                    ore_max_var.set(info.get('ore_max', 48))
                    straordinario_var.set(info.get('straordinario', False))
                    nome_entry.config(state='readonly') # Non permette modifica nome
            else:
                # Pulisci i campi se nessuno è selezionato
                nome_var.set("")
                ore_var.set(40)
                ore_max_var.set(48)
                straordinario_var.set(False)
                nome_entry.config(state='normal')

        lista_addetti.bind('<<ListboxSelect>>', carica_dati_selezionato)

        def salva_addetto():
            nome = nome_var.get().strip()
            if not nome:
                 messagebox.showerror("Errore", "Il nome non può essere vuoto.")
                 return
            if not nome_entry.cget('state') == 'readonly' and nome in self.addetti:
                 messagebox.showerror("Errore", f"L'addetto '{nome}' esiste già.")
                 return

            try:
                 ore_contratto = ore_var.get()
                 ore_max = ore_max_var.get()
                 if ore_contratto < 0 or ore_max < 0:
                     raise ValueError("Le ore non possono essere negative.")
                 if ore_contratto > ore_max:
                     messagebox.showwarning("Attenzione", "Le ore contratto sono maggiori delle ore massime.")
                     # Permetti comunque il salvataggio, ma avvisa
            except (tk.TclError, ValueError) as e:
                 messagebox.showerror("Errore", f"Valore ore non valido: {e}")
                 return

            # Mantiene ferie e riposi esistenti se si sta modificando
            ferie_esistenti = self.addetti.get(nome, {}).get('ferie', [])
            riposi_esistenti = self.addetti.get(nome, {}).get('giorni_riposo', [])

            self.addetti[nome] = {
                'ore_contratto': ore_contratto,
                'ore_max': ore_max,
                'straordinario': straordinario_var.get(),
                'ferie': ferie_esistenti,
                'giorni_riposo': riposi_esistenti
            }

            # Aggiorna lista
            if nome_entry.cget('state') == 'readonly': # Modifica
                 messagebox.showinfo("Successo", f"Dati addetto {nome} aggiornati.")
            else: # Nuovo inserimento
                 lista_addetti.insert(tk.END, nome)
                 # Riordina la lista
                 items = list(lista_addetti.get(0, tk.END))
                 items.sort()
                 lista_addetti.delete(0, tk.END)
                 for item in items:
                      lista_addetti.insert(tk.END, item)
                 messagebox.showinfo("Successo", f"Addetto {nome} aggiunto.")

            self.salva_dati()
            carica_dati_selezionato() # Ricarica per mostrare i dati salvati


        def elimina_addetto():
            selection = lista_addetti.curselection()
            if selection:
                nome = lista_addetti.get(selection[0])
                if messagebox.askyesno("Conferma", f"Vuoi eliminare l'addetto '{nome}'? Verranno perse anche le sue ferie e riposi.", icon='warning'):
                    del self.addetti[nome]
                    lista_addetti.delete(selection[0])
                    self.salva_dati()
                    # Pulisci form dopo eliminazione
                    nome_var.set("")
                    ore_var.set(40)
                    ore_max_var.set(48)
                    straordinario_var.set(False)
                    nome_entry.config(state='normal')
                    messagebox.showinfo("Eliminato", f"Addetto {nome} eliminato.")
            else:
                 messagebox.showwarning("Attenzione", "Selezionare un addetto da eliminare.")

        def nuovo_addetto():
             # Pulisci form e abilita campo nome
             lista_addetti.selection_clear(0, tk.END) # Deseleziona lista
             nome_var.set("")
             ore_var.set(40)
             ore_max_var.set(48)
             straordinario_var.set(False)
             nome_entry.config(state='normal')
             nome_entry.focus() # Metti focus sul campo nome

        # Frame per i bottoni
        frame_bottoni = ttk.Frame(frame_form)
        frame_bottoni.grid(row=4, column=0, columnspan=2, pady=20)

        ttk.Button(frame_bottoni, text="Salva", command=salva_addetto).grid(row=0, column=0, padx=5)
        ttk.Button(frame_bottoni, text="Nuovo", command=nuovo_addetto).grid(row=0, column=1, padx=5)
        ttk.Button(frame_bottoni, text="Elimina Selezionato", command=elimina_addetto).grid(row=0, column=2, padx=5)

        # Forza espansione colonna 1 del frame form
        frame_form.grid_columnconfigure(1, weight=1)


    def gestione_turni(self):
        """Gestisce la definizione dei turni disponibili"""
        # (Codice della funzione gestione_turni originale)
        # ... (incolla qui il codice della funzione gestione_turni dal file originale) ...
        window = tk.Toplevel(self.root)
        window.title("Gestione Turni Disponibili")
        window.geometry("550x400")

        # Frame per la lista dei turni esistenti
        frame_lista = ttk.LabelFrame(window, text="Turni Definiti", padding=10)
        frame_lista.grid(row=0, column=0, padx=10, pady=10, sticky='ns')

        lista_turni = tk.Listbox(frame_lista, width=30, height=10)
        lista_turni.pack(side=tk.LEFT, fill=tk.Y)
        scrollbar = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=lista_turni.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        lista_turni.config(yscrollcommand=scrollbar.set)

        def aggiorna_lista_turni():
            lista_turni.delete(0, tk.END)
            # Ordina i turni per ora di inizio
            turni_ordinati = sorted(self.turni_disponibili, key=lambda t: self._get_orario_in_minuti(t[0]))
            for turno in turni_ordinati:
                lista_turni.insert(tk.END, f"{turno[0]} - {turno[1]}")

        aggiorna_lista_turni() # Carica all'inizio

        # Frame per il form di inserimento nuovo turno
        frame_form = ttk.LabelFrame(window, text="Aggiungi Nuovo Turno", padding=10)
        frame_form.grid(row=0, column=1, padx=10, pady=10, sticky='nsew')

        # Configura grid per espansione
        window.grid_columnconfigure(1, weight=1)
        window.grid_rowconfigure(0, weight=1)

        ttk.Label(frame_form, text="Ora Inizio (HH:MM):").grid(row=0, column=0, pady=5, sticky='w')
        inizio_var = tk.StringVar(value="08:00")
        ttk.Entry(frame_form, textvariable=inizio_var, width=7).grid(row=0, column=1, pady=5, sticky='w')

        ttk.Label(frame_form, text="Ora Fine (HH:MM):").grid(row=1, column=0, pady=5, sticky='w')
        fine_var = tk.StringVar(value="14:00")
        ttk.Entry(frame_form, textvariable=fine_var, width=7).grid(row=1, column=1, pady=5, sticky='w')

        def valida_orario_str(orario_str):
            """Valida il formato e la logica dell'orario stringa HH:MM."""
            try:
                datetime.strptime(orario_str, '%H:%M')
                return True
            except ValueError:
                return False

        def salva_turno():
            inizio = inizio_var.get()
            fine = fine_var.get()

            if not (valida_orario_str(inizio) and valida_orario_str(fine)):
                messagebox.showerror("Errore Formato", "Formato orario non valido. Usare HH:MM (es. 08:00, 14:30).")
                return

            inizio_min = self._get_orario_in_minuti(inizio)
            fine_min = self._get_orario_in_minuti(fine)
            apertura_min = self._get_orario_in_minuti(self.orario_apertura)
            chiusura_min = self._get_orario_in_minuti(self.orario_chiusura)

            if inizio_min >= fine_min:
                messagebox.showerror("Errore Logico", "L'ora di inizio deve essere precedente all'ora di fine.")
                return

            if inizio_min < apertura_min or fine_min > chiusura_min:
                messagebox.showerror("Errore Orario Negozio", f"Il turno deve essere compreso nell'orario di apertura ({self.orario_apertura} - {self.orario_chiusura}).")
                return

            nuovo_turno = (inizio, fine)
            if nuovo_turno in self.turni_disponibili:
                 messagebox.showwarning("Attenzione", f"Il turno {inizio}-{fine} è già presente.")
                 return

            self.turni_disponibili.append(nuovo_turno)
            aggiorna_lista_turni() # Aggiorna la lista visualizzata
            self.salva_dati()
            messagebox.showinfo("Successo", f"Turno {inizio}-{fine} aggiunto correttamente.")
            # Pulisci i campi dopo aggiunta
            inizio_var.set("08:00")
            fine_var.set("14:00")


        def elimina_turno():
            selection = lista_turni.curselection()
            if selection:
                turno_str = lista_turni.get(selection[0])
                inizio, fine = turno_str.split(' - ')
                turno_da_elim = (inizio, fine)

                if messagebox.askyesno("Conferma", f"Vuoi eliminare il turno '{turno_str}'?", icon='warning'):
                    try:
                        self.turni_disponibili.remove(turno_da_elim)
                        aggiorna_lista_turni()
                        self.salva_dati()
                        messagebox.showinfo("Eliminato", f"Turno {turno_str} eliminato.")
                    except ValueError:
                         messagebox.showerror("Errore", "Turno non trovato nei dati interni.") # Non dovrebbe succedere
            else:
                 messagebox.showwarning("Attenzione", "Selezionare un turno da eliminare dalla lista.")


        # Frame per i bottoni
        frame_bottoni = ttk.Frame(frame_form)
        frame_bottoni.grid(row=2, column=0, columnspan=2, pady=20)

        ttk.Button(frame_bottoni, text="Aggiungi Turno", command=salva_turno).grid(row=0, column=0, padx=10)
        ttk.Button(frame_bottoni, text="Elimina Selezionato", command=elimina_turno).grid(row=0, column=1, padx=10)


    def gestione_ferie_riposi(self):
        """Gestisce ferie e giorni di riposo degli addetti"""
        # (Codice della funzione gestione_ferie_riposi originale)
        # ... (incolla qui il codice della funzione gestione_ferie_riposi dal file originale) ...
        if not self.addetti:
            messagebox.showerror("Errore", "Nessun addetto definito. Inserire prima gli addetti.")
            return

        window = tk.Toplevel(self.root)
        window.title("Gestione Ferie e Riposi Settimanali")
        window.geometry("850x650") # Aumentata leggermente

        # Frame principale diviso in due colonne
        frame_sx = ttk.Frame(window, padding=10)
        frame_dx = ttk.Frame(window, padding=10)
        frame_sx.grid(row=0, column=0, sticky='nsew')
        frame_dx.grid(row=0, column=1, sticky='nsew')

        # Configurazione del grid della finestra principale
        window.grid_columnconfigure(0, weight=1)
        window.grid_columnconfigure(1, weight=1)
        window.grid_rowconfigure(0, weight=1)

        # ---- Sezione Selezione Addetto (comune in alto a sinistra) ----
        frame_select = ttk.LabelFrame(frame_sx, text="Selezione Addetto", padding=10)
        frame_select.pack(fill='x', pady=5)

        ttk.Label(frame_select, text="Seleziona Addetto:").pack(side=tk.LEFT, padx=5)
        addetto_var = tk.StringVar()
        # Ordina la lista degli addetti nel combobox
        lista_nomi_addetti = sorted(list(self.addetti.keys()))
        combo_addetti = ttk.Combobox(frame_select, textvariable=addetto_var,
                                   values=lista_nomi_addetti, state='readonly', width=30)
        combo_addetti.pack(side=tk.LEFT, padx=5, fill='x', expand=True)


        # ---- Sezione Gestione Ferie (sotto selezione addetto, sinistra) ----
        frame_ferie = ttk.LabelFrame(frame_sx, text="Gestione Ferie (Seleziona Giorni)", padding=10)
        frame_ferie.pack(fill='both', expand=True, pady=10)

        # Calendario per selezione ferie
        frame_nav_calendario = ttk.Frame(frame_ferie)
        frame_nav_calendario.pack(pady=5, fill='x')

        # Selezione anno e mese
        current_year = datetime.now().year
        current_month = datetime.now().month
        anno_var = tk.IntVar(value=current_year)
        mese_var = tk.IntVar(value=current_month)

        def update_mese_anno_label():
             anno = anno_var.get()
             mese = mese_var.get()
             mese_anno_label.config(text=f"{calendar.month_name[mese]} {anno}")

        def prev_month():
            mese = mese_var.get()
            anno = anno_var.get()
            mese -= 1
            if mese == 0:
                mese = 12
                anno -= 1
            mese_var.set(mese)
            anno_var.set(anno)
            update_mese_anno_label()
            aggiorna_calendario_ferie()

        def next_month():
            mese = mese_var.get()
            anno = anno_var.get()
            mese += 1
            if mese == 13:
                mese = 1
                anno += 1
            mese_var.set(mese)
            anno_var.set(anno)
            update_mese_anno_label()
            aggiorna_calendario_ferie()

        ttk.Button(frame_nav_calendario, text="<", command=prev_month, width=3).pack(side=tk.LEFT, padx=5)
        mese_anno_label = ttk.Label(frame_nav_calendario, text="", font=('Helvetica', 12, 'bold'), width=20, anchor='center')
        mese_anno_label.pack(side=tk.LEFT, padx=5, fill='x', expand=True)
        ttk.Button(frame_nav_calendario, text=">", command=next_month, width=3).pack(side=tk.LEFT, padx=5)

        # Frame per i giorni del calendario
        frame_giorni_cal = ttk.Frame(frame_ferie)
        frame_giorni_cal.pack(pady=10, fill='both', expand=True)
        giorni_checkbox_vars = {} # Dizionario {giorno_num: tk.BooleanVar}

        # ----- Lista Ferie Programmate (sotto calendario, sinistra) ------
        frame_lista_ferie = ttk.LabelFrame(frame_sx, text="Ferie Programmate per Addetto", padding=10)
        frame_lista_ferie.pack(fill='x', pady=5)

        lista_ferie_display = tk.Listbox(frame_lista_ferie, height=6)
        scrollbar_ferie = ttk.Scrollbar(frame_lista_ferie, orient=tk.VERTICAL, command=lista_ferie_display.yview)
        lista_ferie_display.config(yscrollcommand=scrollbar_ferie.set)
        lista_ferie_display.pack(side=tk.LEFT, fill='both', expand=True)
        scrollbar_ferie.pack(side=tk.RIGHT, fill='y')


        def aggiorna_calendario_ferie():
            """Aggiorna la visualizzazione del calendario delle ferie."""
            # Pulisci frame giorni calendario
            for widget in frame_giorni_cal.winfo_children():
                widget.destroy()
            giorni_checkbox_vars.clear()

            try:
                anno = anno_var.get()
                mese = mese_var.get()
            except (tk.TclError, ValueError):
                messagebox.showerror("Errore", "Anno o mese non valido.")
                return

            # Crea intestazioni giorni settimana
            giorni_settimana_short = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
            for i, giorno_abbr in enumerate(giorni_settimana_short):
                lbl = ttk.Label(frame_giorni_cal, text=giorno_abbr, width=5, anchor='center', relief="groove")
                lbl.grid(row=0, column=i, padx=1, pady=1, sticky='nsew')

            # Ottieni il calendario del mese (matrice settimane x giorni)
            try:
                cal = calendar.monthcalendar(anno, mese)
            except ValueError:
                 messagebox.showerror("Errore", f"Data non valida: {mese}/{anno}")
                 return

            # Crea checkbutton per ogni giorno del mese selezionato
            addetto_selezionato = addetto_var.get()
            ferie_addetto = set() # Usiamo un set per controlli veloci
            if addetto_selezionato and addetto_selezionato in self.addetti:
                 ferie_addetto = set(self.addetti[addetto_selezionato].get('ferie', []))

            for riga, settimana in enumerate(cal):
                for col, giorno_num in enumerate(settimana):
                    if giorno_num != 0:
                        data_corrente = datetime(anno, mese, giorno_num)
                        data_str = data_corrente.strftime('%Y-%m-%d')

                        var = tk.BooleanVar()
                        # Seleziona la checkbox se il giorno è nelle ferie dell'addetto
                        var.set(data_str in ferie_addetto)

                        giorni_checkbox_vars[giorno_num] = var
                        cb = ttk.Checkbutton(frame_giorni_cal, text=str(giorno_num),
                                           variable=var, width=4) # style='Toolbutton' potrebbe essere carino
                        cb.grid(row=riga + 1, column=col, padx=1, pady=1, sticky='nsew')
                    else:
                         # Cella vuota per giorni fuori dal mese
                         ttk.Label(frame_giorni_cal, text="").grid(row=riga + 1, column=col, padx=1, pady=1)

            # Configura le colonne del grid per avere uguale larghezza
            for i in range(7):
                frame_giorni_cal.grid_columnconfigure(i, weight=1)


        def salva_ferie_selezionate():
            """Salva le ferie selezionate nel calendario per l'addetto corrente."""
            addetto = addetto_var.get()
            if not addetto:
                messagebox.showerror("Errore", "Nessun addetto selezionato.")
                return

            try:
                anno = anno_var.get()
                mese = mese_var.get()
            except (tk.TclError, ValueError):
                 messagebox.showerror("Errore", "Anno o mese non valido per il salvataggio.")
                 return

            # Prendi le ferie esistenti dell'addetto (anche di altri mesi/anni)
            ferie_attuali = set(self.addetti[addetto].get('ferie', []))

            # Rimuovi le date del mese corrente dal set esistente
            ferie_attuali = {data for data in ferie_attuali if not data.startswith(f"{anno:04d}-{mese:02d}-")}

            # Aggiungi le date selezionate nel calendario corrente
            for giorno, var in giorni_checkbox_vars.items():
                if var.get():
                    data_str = f"{anno:04d}-{mese:02d}-{giorno:02d}"
                    ferie_attuali.add(data_str)

            # Aggiorna i dati dell'addetto e salva
            self.addetti[addetto]['ferie'] = sorted(list(ferie_attuali)) # Salva come lista ordinata
            self.salva_dati()
            aggiorna_lista_ferie_display() # Aggiorna la listbox sotto
            messagebox.showinfo("Successo", f"Ferie per {addetto} aggiornate correttamente.")


        ttk.Button(frame_ferie, text="Salva Ferie per Mese Corrente",
                  command=salva_ferie_selezionate).pack(pady=10)

        # ---- Sezione Gestione Riposi (destra) ----
        frame_riposi = ttk.LabelFrame(frame_dx, text="Gestione Giorni di Riposo Settimanali Fissi", padding=10)
        frame_riposi.pack(fill='both', expand=True, pady=5)

        # Checkbox per ogni giorno della settimana
        giorni_settimana_completi = ["Lunedì", "Martedì", "Mercoledì", "Giovedì",
                                   "Venerdì", "Sabato", "Domenica"]
        riposi_checkbox_vars = [] # Lista di tk.BooleanVar

        for i, giorno_nome in enumerate(giorni_settimana_completi):
            var = tk.BooleanVar()
            riposi_checkbox_vars.append(var)
            cb = ttk.Checkbutton(frame_riposi, text=giorno_nome, variable=var)
            cb.pack(anchor='w', padx=10, pady=2)

        # Lista dei riposi attualmente impostati per l'addetto
        ttk.Label(frame_riposi, text="\nRiposi Settimanali Attuali:").pack(pady=5)
        lista_riposi_display = tk.Listbox(frame_riposi, height=7) # 7 giorni max
        lista_riposi_display.pack(fill='x', padx=10, pady=5)


        def salva_giorni_riposo():
            """Salva i giorni di riposo settimanali selezionati per l'addetto."""
            addetto = addetto_var.get()
            if not addetto:
                messagebox.showerror("Errore", "Nessun addetto selezionato.")
                return

            giorni_riposo_indices = [i for i, var in enumerate(riposi_checkbox_vars) if var.get()]

            # if not giorni_riposo_indices: # Permettiamo zero giorni di riposo? Sì.
            #     messagebox.showwarning("Attenzione", "Nessun giorno di riposo selezionato.")
                # return

            self.addetti[addetto]['giorni_riposo'] = giorni_riposo_indices
            self.salva_dati()
            aggiorna_lista_riposi_display() # Aggiorna la listbox
            messagebox.showinfo("Successo", f"Giorni di riposo per {addetto} salvati.")


        ttk.Button(frame_riposi, text="Salva Giorni di Riposo",
                  command=salva_giorni_riposo).pack(pady=20)


        # --- Funzioni di Aggiornamento Comuni ---
        def aggiorna_lista_ferie_display():
            """Aggiorna la listbox che mostra le ferie programmate."""
            lista_ferie_display.delete(0, tk.END)
            addetto = addetto_var.get()
            if addetto and addetto in self.addetti:
                # Mostra solo le ferie da oggi in poi, ordinate
                oggi_str = datetime.now().strftime('%Y-%m-%d')
                ferie_future = sorted([d for d in self.addetti[addetto].get('ferie', []) if d >= oggi_str])
                if ferie_future:
                    for data_str in ferie_future:
                        try:
                             data_dt = datetime.strptime(data_str, '%Y-%m-%d')
                             lista_ferie_display.insert(tk.END, data_dt.strftime('%d %b %Y (%a)')) # Formato leggibile
                        except ValueError:
                             lista_ferie_display.insert(tk.END, data_str + " (Formato errato?)") # Fallback
                else:
                    lista_ferie_display.insert(tk.END, "Nessuna feria futura programmata.")


        def aggiorna_lista_riposi_display():
            """Aggiorna la listbox che mostra i giorni di riposo settimanali."""
            lista_riposi_display.delete(0, tk.END)
            addetto = addetto_var.get()
            if addetto and addetto in self.addetti:
                riposi_indices = sorted(self.addetti[addetto].get('giorni_riposo', []))
                if riposi_indices:
                     for index in riposi_indices:
                          if 0 <= index < len(giorni_settimana_completi):
                               lista_riposi_display.insert(tk.END, giorni_settimana_completi[index])
                else:
                    lista_riposi_display.insert(tk.END, "Nessun giorno di riposo fisso.")


        # Callback quando cambia l'addetto selezionato
        def on_select_addetto(event=None):
            """Aggiorna tutti i display quando viene selezionato un nuovo addetto."""
            addetto = addetto_var.get()
            if addetto and addetto in self.addetti:
                # Aggiorna Checkbox Riposi
                riposi_attuali = self.addetti[addetto].get('giorni_riposo', [])
                for i, var in enumerate(riposi_checkbox_vars):
                    var.set(i in riposi_attuali)

                # Aggiorna Lista Riposi Display
                aggiorna_lista_riposi_display()

                # Aggiorna Calendario Ferie (mostrando quelle dell'addetto)
                aggiorna_calendario_ferie()

                # Aggiorna Lista Ferie Display
                aggiorna_lista_ferie_display()

            else: # Se nessun addetto è selezionato o non valido
                 # Pulisci tutto
                 for var in riposi_checkbox_vars: var.set(False)
                 aggiorna_lista_riposi_display()
                 aggiorna_calendario_ferie() # Pulisce le selezioni nel calendario
                 aggiorna_lista_ferie_display()


        # Binding evento e inizializzazione
        combo_addetti.bind('<<ComboboxSelected>>', on_select_addetto)
        update_mese_anno_label() # Imposta etichetta mese/anno iniziale
        # Se c'è un addetto selezionato all'inizio (primo della lista), carica i suoi dati
        if lista_nomi_addetti:
             addetto_var.set(lista_nomi_addetti[0])
             on_select_addetto()
        else:
             # Se non ci sono addetti, svuota/inizializza i campi
             on_select_addetto()


    # ==========================================================================
    #               NUOVA LOGICA DI PIANIFICAZIONE REFACTORED
    # ==========================================================================

    # --- Funzioni di Utilità per Orari ---
    def _get_orario_in_minuti(self, orario_str):
        """Converte una stringa orario HH:MM in minuti da mezzanotte."""
        try:
            ore, minuti = map(int, orario_str.split(':'))
            return ore * 60 + minuti
        except ValueError:
            # Gestisce orari potenzialmente invalidi o formati diversi
            print(f"Attenzione: formato orario non valido '{orario_str}'")
            return None # O solleva un'eccezione specifica

    def _get_orario_da_minuti(self, minuti_totali):
        """Converte minuti da mezzanotte in una stringa orario HH:MM."""
        if minuti_totali is None:
            return "N/A"
        ore = (minuti_totali // 60) % 24
        minuti = minuti_totali % 60
        return f"{ore:02d}:{minuti:02d}"

    # --- Funzioni Helper per la Pianificazione Refactored ---
    # --- Funzione Helper per Calcolare Festività ---
    def _get_festivi_mese(self, anno, mese):
        """
        Calcola l'insieme delle festività (formato 'dd-mm') per un dato anno e mese.
        Include festività fisse e Pasqua/Pasquetta.
        """
        # Funzione interna per Pasqua (così non la ripetiamo)
        def _calcola_pasqua(year):
            # Algoritmo di Meeus/Jones/Butcher (Gregoriano)
            a = year % 19; b = year // 100; c = year % 100; d = b // 4; e = b % 4
            f = (b + 8) // 25; g = (b - f + 1) // 3; h = (19 * a + b - d - g + 15) % 30
            i = c // 4; k = c % 4; l = (32 + 2 * e + 2 * i - h - k) % 7
            m = (a + 11 * h + 22 * l) // 451
            month = (h + l - 7 * m + 114) // 31
            day = ((h + l - 7 * m + 114) % 31) + 1
            return datetime(year, month, day)

        try:
            pasqua = _calcola_pasqua(anno)
            pasquetta = pasqua + timedelta(days=1)
            # Considera solo se Pasqua/Pasquetta cadono nel mese richiesto?
            # Per ora le includiamo sempre nel set annuale.
            festivita_calcolate = {pasqua.strftime('%d-%m'), pasquetta.strftime('%d-%m')}
        except Exception as e_pasqua:
            print(f"Errore nel calcolo di Pasqua per l'anno {anno}: {e_pasqua}")
            festivita_calcolate = set()

        # Unisci festività fisse (da self.giorni_festivi) e calcolate
        festivi_anno_corrente = set(self.giorni_festivi) | festivita_calcolate
        return festivi_anno_corrente

    def _trova_addetti_disponibili_giorno(self, data):
        """
        Restituisce una lista di nomi di addetti disponibili per una data specifica.
        Controlla solo ferie e giorno di riposo settimanale.
        """
        addetti_disponibili = []
        data_str_ymd = data.strftime('%Y-%m-%d')
        giorno_settimana = data.weekday() # Lunedì è 0, Domenica è 6

        for nome, info in self.addetti.items():
            # Controlla ferie
            if data_str_ymd in info.get('ferie', []):
                continue  # In ferie

            # Controlla giorni di riposo
            if giorno_settimana in info.get('giorni_riposo', []):
                continue  # Giorno di riposo settimanale

            addetti_disponibili.append(nome)

        return addetti_disponibili

    def _calcola_ore_lavorate_mese(self, addetto, mese_calendario):
        """
        Calcola le ore totali lavorate da un addetto fino al giorno prima
        nel calendario parziale del mese corrente.
        """
        ore_totali = 0
        for giorno, turni_giorno in mese_calendario.items():
            # Controlla se l'addetto ha un turno assegnato (non 'RIPOSO', 'FERIE', etc.)
            turno_info = turni_giorno.get(addetto)
            if isinstance(turno_info, (list, tuple)) and len(turno_info) == 2:
                turno_orari = turno_info # È una tupla ('HH:MM', 'HH:MM')
                inizio_min = self._get_orario_in_minuti(turno_orari[0])
                fine_min = self._get_orario_in_minuti(turno_orari[1])
                if inizio_min is not None and fine_min is not None:
                    durata_min = (fine_min - inizio_min)
                    if durata_min < 0: # Gestione mezzanotte (improbabile)
                        durata_min += 24 * 60
                    ore_totali += durata_min / 60.0
        return ore_totali


    def _verifica_vincoli_turno(self, addetto, turno, data, ore_lavorate_mese_corrente):
        """
        Verifica i vincoli *rigidi* per assegnare un turno a un addetto in una data.
        Restituisce True se i vincoli sono rispettati, False altrimenti.
        """
        info_addetto = self.addetti[addetto]
        ore_max = info_addetto.get('ore_max', 48) # Default a 48 se non specificato
        permette_straordinario = info_addetto.get('straordinario', False)

        inizio_min = self._get_orario_in_minuti(turno[0])
        fine_min = self._get_orario_in_minuti(turno[1])

        if inizio_min is None or fine_min is None:
            return False # Turno non valido

        durata_min = (fine_min - inizio_min)
        if durata_min < 0: durata_min += 24 * 60 # Gestione mezzanotte (improbabile qui)
        ore_turno = durata_min / 60.0

        # 1. Vincolo Ore Massime (solo se non permette straordinario)
        #    Usiamo una piccola tolleranza per evitare problemi di floating point
        if not permette_straordinario:
            if (ore_lavorate_mese_corrente + ore_turno) > (ore_max + 0.01):
                # print(f"Vincolo violato: {addetto} supererebbe ore max ({ore_lavorate_mese_corrente + ore_turno:.1f} > {ore_max})")
                return False

        # 2. Vincolo Riposo Minimo tra Turni (Esempio: almeno 11 ore)
        #    TODO: Implementare questo controllo guardando il turno del giorno precedente
        #    Per ora ritorna sempre True su questo vincolo

        return True # Tutti i vincoli rigidi verificati (per ora)


    def _calcola_punteggio_turno_refactored(self, addetto, turno, data, ore_lavorate_mese_corrente, mese_calendario):
        """
        Calcola un punteggio di "desiderabilità" per un'assegnazione valida.
        Punteggi più alti sono migliori. Qui implementiamo una logica semplice.
        """
        punteggio = 100 # Punteggio base

        info_addetto = self.addetti[addetto]
        ore_contratto = info_addetto.get('ore_contratto', 40)

        inizio_min = self._get_orario_in_minuti(turno[0])
        fine_min = self._get_orario_in_minuti(turno[1])
        if inizio_min is None or fine_min is None: return -1000 # Turno invalido
        durata_min = (fine_min - inizio_min)
        if durata_min < 0: durata_min += 24 * 60
        ore_turno = durata_min / 60.0

        # 1. Bonus se sotto le ore contratto
        if ore_lavorate_mese_corrente + ore_turno <= ore_contratto:
            punteggio += 20

        # 2. Malus per turni uguali recenti (guardiamo ultimi 3 giorni nel calendario parziale)
        turno_str = f"{turno[0]}-{turno[1]}"
        giorno_corrente = data.day
        turni_recenti_uguali = 0
        for d in range(max(1, giorno_corrente - 3), giorno_corrente):
            if d in mese_calendario:
                turno_passato_info = mese_calendario[d].get(addetto)
                if isinstance(turno_passato_info, (list, tuple)) and len(turno_passato_info) == 2:
                     if f"{turno_passato_info[0]}-{turno_passato_info[1]}" == turno_str:
                         turni_recenti_uguali += 1
        punteggio -= turni_recenti_uguali * 30 # Penalità crescente

        # 3. Bonus per alternanza mattina/pomeriggio (molto semplificato)
        #    Contiamo turni mattina/pomeriggio nel mese finora
        mattina = 0
        pomeriggio = 0
        for d, turni_giorno in mese_calendario.items():
             turno_passato_info = turni_giorno.get(addetto)
             if isinstance(turno_passato_info, (list, tuple)) and len(turno_passato_info) == 2:
                orario_inizio_passato = self._get_orario_in_minuti(turno_passato_info[0])
                if orario_inizio_passato is not None and orario_inizio_passato < self._get_orario_in_minuti("13:00"):
                   mattina += 1
                else:
                   pomeriggio += 1

        ora_inizio_turno_corrente = self._get_orario_in_minuti(turno[0])
        if ora_inizio_turno_corrente is not None:
             if ora_inizio_turno_corrente < self._get_orario_in_minuti("13:00"): # È un turno di mattina
                 if mattina <= pomeriggio: # Favorisce se ha fatto meno mattine
                     punteggio += 10
             else: # È un turno di pomeriggio
                 if pomeriggio < mattina: # Favorisce se ha fatto meno pomeriggi
                     punteggio += 10

        return punteggio

    def _seleziona_turni_giornalieri(self, data, addetti_disponibili, mese_calendario):
        """
        Seleziona la migliore combinazione di turni per coprire l'orario 8:00-21:00,
        dando priorità assoluta alla copertura.
        Utilizza una strategia greedy focalizzata sulla copertura.
        """
        turni_assegnati_giorno = {} # {nome_addetto: ('HH:MM', 'HH:MM'), ...}
        orario_inizio_min = self._get_orario_in_minuti(self.orario_apertura)
        orario_fine_min = self._get_orario_in_minuti(self.orario_chiusura)

        if orario_inizio_min is None or orario_fine_min is None:
             print(f"Errore: Orari di apertura/chiusura non validi ({self.orario_apertura}-{self.orario_chiusura})")
             return {'ERRORE': 'Orari negozio non validi'}

        # Array booleano per tracciare la copertura minuto per minuto
        minuti_da_coprire = orario_fine_min - orario_inizio_min
        if minuti_da_coprire <= 0:
            print(f"Errore: Orario di chiusura ({self.orario_chiusura}) non successivo all'apertura ({self.orario_apertura})")
            return {'ERRORE': 'Orario negozio illogico'}

        copertura_minuti = [False] * minuti_da_coprire # False = scoperto, True = coperto

        # Lista degli addetti ancora assegnabili oggi
        addetti_non_assegnati = addetti_disponibili.copy()

        # 1. Genera tutte le possibili assegnazioni VALIDE per oggi
        assegnazioni_possibili = []
        for addetto in addetti_disponibili:
            # Calcola le ore lavorate finora nel mese per questo addetto
            ore_lavorate_mese = self._calcola_ore_lavorate_mese(addetto, mese_calendario)
            for turno in self.turni_disponibili:
                if self._verifica_vincoli_turno(addetto, turno, data, ore_lavorate_mese):
                    punteggio = self._calcola_punteggio_turno_refactored(addetto, turno, data, ore_lavorate_mese, mese_calendario)
                    inizio_min = self._get_orario_in_minuti(turno[0])
                    fine_min = self._get_orario_in_minuti(turno[1])
                    # Aggiungi solo se gli orari sono validi
                    if inizio_min is not None and fine_min is not None:
                         assegnazioni_possibili.append({
                             'addetto': addetto,
                             'turno': turno,
                             'punteggio': punteggio,
                             'inizio_min': inizio_min,
                             'fine_min': fine_min
                         })

        # Ordina le possibilità: Prima per punteggio (più alto è meglio), poi per durata (più lungo è meglio per copertura)
        assegnazioni_possibili.sort(key=lambda x: (x['punteggio'], x['fine_min'] - x['inizio_min']), reverse=True)

        # 2. Ciclo Greedy per Copertura: continua finché c'è qualcosa da coprire e ci sono opzioni
        minuti_coperti_count = 0
        while minuti_coperti_count < minuti_da_coprire:
            prima_ora_scoperta_idx = -1 # Indice nell'array copertura_minuti
            for i in range(minuti_da_coprire):
                if not copertura_minuti[i]:
                    prima_ora_scoperta_idx = i
                    break

            if prima_ora_scoperta_idx == -1:
                # print(f"   Giorno {data.day}: Copertura completata.")
                break # Tutto coperto!

            prima_ora_scoperta_min = orario_inizio_min + prima_ora_scoperta_idx

            # Trova la migliore assegnazione possibile che:
            # - Copre la `prima_ora_scoperta_min`
            # - Usa un addetto non ancora assegnato oggi
            # - Ha il punteggio/durata migliore tra quelle che coprono l'ora
            migliore_assegnazione_per_gap = None
            indice_da_processare = -1 # Indice nell'array assegnazioni_possibili

            for i, ass in enumerate(assegnazioni_possibili):
                # Controlla se l'addetto è disponibile e se il turno copre l'ora scoperta
                if (ass['addetto'] in addetti_non_assegnati and
                    ass['inizio_min'] <= prima_ora_scoperta_min < ass['fine_min']):

                    migliore_assegnazione_per_gap = ass
                    indice_da_processare = i
                    break # Trovata la migliore disponibile secondo l'ordinamento iniziale

            # Se non abbiamo trovato NESSUNA assegnazione per coprire il buco
            if migliore_assegnazione_per_gap is None:
                print(f"   Attenzione Giorno {data.day}: Impossibile trovare turno valido per coprire ora {self._get_orario_da_minuti(prima_ora_scoperta_min)}. Copertura parziale.")
                # Assegna 'ERRORE' a tutti gli addetti disponibili non assegnati per segnalare? No, lasciamo vuoto.
                turni_assegnati_giorno['ERRORE_COPERTURA'] = f"Buco dalle {self._get_orario_da_minuti(prima_ora_scoperta_min)}"
                break # Interrompi il ciclo, non si può coprire oltre

            # Assegna il turno trovato
            addetto_scelto = migliore_assegnazione_per_gap['addetto']
            turno_scelto = migliore_assegnazione_per_gap['turno']
            turni_assegnati_giorno[addetto_scelto] = turno_scelto
            # print(f"   Assegnato {addetto_scelto} al turno {turno_scelto} per coprire {self._get_orario_da_minuti(prima_ora_scoperta_min)}")

            # Aggiorna la copertura
            start_idx = max(0, migliore_assegnazione_per_gap['inizio_min'] - orario_inizio_min)
            end_idx = min(minuti_da_coprire, migliore_assegnazione_per_gap['fine_min'] - orario_inizio_min)
            for i in range(start_idx, end_idx):
                if not copertura_minuti[i]:
                    copertura_minuti[i] = True
            minuti_coperti_count = sum(copertura_minuti) # Ricalcola esattamente quanti minuti sono coperti

            # Rimuovi l'addetto da quelli disponibili oggi
            addetti_non_assegnati.remove(addetto_scelto)
            # Rimuovi TUTTE le altre possibili assegnazioni per l'addetto scelto oggi dalla lista
            # per evitare di riassegnarlo per errore (più sicuro che rimuovere solo l'indice)
            assegnazioni_possibili = [a for a in assegnazioni_possibili if a['addetto'] != addetto_scelto]


        # Verifica finale copertura (opzionale, per sicurezza)
        if minuti_coperti_count < minuti_da_coprire:
             if 'ERRORE_COPERTURA' not in turni_assegnati_giorno: # Evita doppioni se già segnalato
                 prima_ora_scoperta_idx = -1
                 for i in range(minuti_da_coprire):
                     if not copertura_minuti[i]:
                         prima_ora_scoperta_idx = i
                         break
                 ora_buco = self._get_orario_da_minuti(orario_inizio_min + prima_ora_scoperta_idx) if prima_ora_scoperta_idx != -1 else "N/D"
                 print(f"   Giorno {data.day}: Copertura INCOMPLETA! {minuti_coperti_count}/{minuti_da_coprire} minuti coperti. Buco da {ora_buco}.")
                 turni_assegnati_giorno['ERRORE_COPERTURA'] = f"Incompleta ({ora_buco})"

        # Aggiungi riposo/ferie/non assegnato per chi non ha lavorato
        for addetto in self.addetti.keys(): # Itera su tutti gli addetti
             if addetto not in turni_assegnati_giorno:
                # Controlla se era in ferie o riposo originale
                data_str_ymd = data.strftime('%Y-%m-%d')
                giorno_settimana = data.weekday()
                info_addetto = self.addetti[addetto]
                if data_str_ymd in info_addetto.get('ferie', []):
                     turni_assegnati_giorno[addetto] = 'FERIE'
                elif giorno_settimana in info_addetto.get('giorni_riposo', []):
                     turni_assegnati_giorno[addetto] = 'RIPOSO'
                # Altrimenti, se era disponibile ma non assegnato, non mettere nulla o 'NON ASSEGNATO'
                # else:
                #    turni_assegnati_giorno[addetto] = '-' # O lascia vuoto

        return turni_assegnati_giorno


    # --- Funzione Principale Refactored ---
    # --- Funzione Principale Refactored (Sostituisce l'originale) ---
    def _genera_calendario_mensile_refactored(self, anno, mese):
        """
        Genera il calendario mensile dando priorità alla copertura oraria completa
        e utilizzando funzioni helper per separare le logiche.
        """
        # Ottieni le festività per l'anno corrente usando la funzione helper
        festivi_anno_corrente = self._get_festivi_mese(anno, mese) # Passiamo anche il mese, anche se non usato per ora dalla helper

        num_giorni = calendar.monthrange(anno, mese)[1]
        calendario_mensile = {} # {1: {nome: turno/stato, ...}, 2: {...}}

        # Prova a ottenere il nome del mese in italiano
        try:
            nome_mese_locale = calendar.month_name[mese]
        except IndexError:
            nome_mese_locale = f"Mese {mese}"

        print(f"\n--- Generazione Pianificazione per {nome_mese_locale} {anno} ---")
        print(f"Festività considerate (formato gg-mm): {', '.join(sorted(list(festivi_anno_corrente)))}")

        for giorno in range(1, num_giorni + 1):
            data = datetime(anno, mese, giorno)
            data_str_dm = data.strftime('%d-%m') # Per controllo festivi
            giorno_settimana_abbr = data.strftime('%a') # Es: Lun, Mar...

            print(f"\n-- Giorno {giorno} ({giorno_settimana_abbr}) --")

            # Salta i giorni festivi
            if data_str_dm in festivi_anno_corrente: # Usa la variabile definita sopra
                print("   Festivo - Saltato")
                # Marca come festivo per tutti, tranne chi è in ferie quel giorno
                calendario_mensile[giorno] = {}
                for nome_addetto, info_addetto in self.addetti.items():
                     if data.strftime('%Y-%m-%d') in info_addetto.get('ferie', []):
                          calendario_mensile[giorno][nome_addetto] = 'FERIE'
                     else:
                          calendario_mensile[giorno][nome_addetto] = 'FESTIVO'
                continue

            # 1. Trova addetti disponibili oggi (considera ferie e riposi settimanali)
            addetti_disponibili_oggi = self._trova_addetti_disponibili_giorno(data)

            if not addetti_disponibili_oggi:
                print("   ATTENZIONE: Nessun addetto disponibile per questo giorno!")
                # Marca tutti come non disponibili (o errore?)
                calendario_mensile[giorno] = {}
                for nome_addetto, info_addetto in self.addetti.items():
                    data_str_ymd = data.strftime('%Y-%m-%d')
                    giorno_settimana = data.weekday()
                    if data_str_ymd in info_addetto.get('ferie', []):
                         calendario_mensile[giorno][nome_addetto] = 'FERIE'
                    elif giorno_settimana in info_addetto.get('giorni_riposo', []):
                         calendario_mensile[giorno][nome_addetto] = 'RIPOSO'
                    else:
                         # Questo caso dovrebbe essere raro se _trova_addetti_disponibili_giorno funziona
                         calendario_mensile[giorno][nome_addetto] = 'ERRORE_NODISP'
                continue

            print(f"   Addetti potenzialmente disponibili: {', '.join(addetti_disponibili_oggi)}")

            # 2. Seleziona i turni per la giornata dando priorità alla copertura
            #    Passa il calendario parziale (mese_calendario) per i controlli sui vincoli
            turni_del_giorno = self._seleziona_turni_giornalieri(data, addetti_disponibili_oggi, calendario_mensile)

            # 3. Aggiungi i turni selezionati al calendario mensile
            #    La funzione _seleziona_turni_giornalieri già include Ferie/Riposo per chi non lavora
            calendario_mensile[giorno] = turni_del_giorno

            # Stampa i turni assegnati per il giorno (debug)
            if turni_del_giorno:
                 print("   Turni/Stato assegnati:")
                 # Ordina per nome addetto per una stampa più leggibile
                 for addetto, stato_turno in sorted(turni_del_giorno.items()):
                     if isinstance(stato_turno, (list, tuple)): # È un turno
                         print(f"     {addetto}: {stato_turno[0]}-{stato_turno[1]}")
                     elif isinstance(stato_turno, str): # È uno stato (FERIE, RIPOSO, ERRORE...)
                         print(f"     {addetto}: {stato_turno}")
            else:
                 # Questo non dovrebbe accadere con la logica attuale, ma è una sicurezza
                 print("   Nessun turno assegnato per il giorno (potrebbe essere errore logico).")


        # Calcolo finale e stampa riepilogo ore (opzionale, ma utile)
        print("\n--- Riepilogo Ore Lavorate Stimate nel Mese ---")
        for addetto, info in sorted(self.addetti.items()): # Ordina per nome
            ore_finali = self._calcola_ore_lavorate_mese(addetto, calendario_mensile)
            ore_contratto = info.get('ore_contratto', 0) # Usa 0 come default se manca
            ore_max = info.get('ore_max', 0) # Usa 0 come default se manca
            stato = "OK"
            # Usiamo tolleranza per confronti float
            if ore_max > 0 and ore_finali > (ore_max + 0.01) and not info.get('straordinario', False):
                stato = f"!!! ERRORE: Superato limite ore max ({ore_max}) di {ore_finali - ore_max:.1f} ore!"
            elif ore_max > 0 and ore_finali > (ore_max + 0.01):
                 stato = f"Straordinario (+{ore_finali - ore_max:.1f} ore)"
            elif ore_contratto > 0 and ore_finali < (ore_contratto - 0.01):
                 stato = f"Sotto contratto ({ore_contratto}) di {ore_contratto - ore_finali:.1f} ore"

            print(f"{addetto}: {ore_finali:.1f} ore (Contr: {ore_contratto}, Max: {ore_max}) - {stato}")

        print("\n--- Fine Generazione Pianificazione ---")
        return calendario_mensile

    def _salva_calendario_excel(self, calendario, anno, mese):
        """Salva il calendario dei turni su file Excel con formattazione migliorata"""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Impostazioni di base del foglio
        try:
            nome_mese = calendar.month_name[mese]
        except IndexError:
            nome_mese = f"Mese {mese}" # Fallback
        ws.title = f"Turni {nome_mese} {anno}"
        ws.sheet_view.zoomScale = 85

        # Stili comuni
        thin_border_side = Side(style='thin', color='A0A0A0')
        bordo_sottile = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side
        )
        allineamento_centro = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        allineamento_sinistra = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )

        # Formattazione header
        header_font = Font(bold=True, size=11, color='000000')
        header_fill = PatternFill(start_color=self.colori['header'],
                                end_color=self.colori['header'],
                                fill_type='solid')

        # Scrivi intestazione (Data e Giorno Sett.)
        ws.cell(1, 1, "Data").font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 1).border = bordo_sottile
        ws.cell(1, 1).alignment = allineamento_centro
        ws.column_dimensions['A'].width = 15

        # Scrivi nomi addetti nelle colonne
        nomi_addetti_ordinati = sorted(self.addetti.keys())
        for col, addetto in enumerate(nomi_addetti_ordinati, 2):
            cell = ws.cell(1, col, addetto)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = bordo_sottile
            cell.alignment = allineamento_centro
            ws.column_dimensions[get_column_letter(col)].width = 18 # Larghezza colonne addetti

        # Calcola Pasqua per festività
        def calcola_pasqua(year):
            a = year % 19; b = year // 100; c = year % 100; d = b // 4; e = b % 4
            f = (b + 8) // 25; g = (b - f + 1) // 3; h = (19 * a + b - d - g + 15) % 30
            i = c // 4; k = c % 4; l = (32 + 2 * e + 2 * i - h - k) % 7
            m = (a + 11 * h + 22 * l) // 451
            month = (h + l - 7 * m + 114) // 31
            day = ((h + l - 7 * m + 114) % 31) + 1
            return datetime(year, month, day)
        pasqua = calcola_pasqua(anno)
        pasquetta = pasqua + timedelta(days=1)
        festivita_calcolate = {pasqua.strftime('%d-%m'), pasquetta.strftime('%d-%m')}
        festivi_mese_corrente = set(self.giorni_festivi) | festivita_calcolate


        # Scrivi i giorni e i turni/stati
        num_giorni_mese = calendar.monthrange(anno, mese)[1]
        for giorno in range(1, num_giorni_mese + 1):
            data = datetime(anno, mese, giorno)
            data_str_dm = data.strftime('%d-%m')
            data_str_ymd = data.strftime('%Y-%m-%d')
            giorno_settimana_abbr = data.strftime('%a') # Es: Lun, Mar...

            # Formattazione riga
            riga = giorno + 1

            # Scrivi data e giorno settimana
            cell_data = ws.cell(riga, 1, f"{giorno:02d}/{mese:02d}/{anno} ({giorno_settimana_abbr})")
            cell_data.border = bordo_sottile
            cell_data.alignment = allineamento_sinistra # Allinea a sinistra per leggibilità

            # Determina colore di sfondo per la riga del giorno
            is_festivo = data_str_dm in festivi_mese_corrente
            is_weekend = data.weekday() >= 5 # Sabato=5, Domenica=6

            fill_giorno = None
            if is_festivo:
                fill_giorno = PatternFill(start_color=self.colori['festivo'], end_color=self.colori['festivo'], fill_type='solid')
            elif is_weekend:
                fill_giorno = PatternFill(start_color=self.colori['weekend'], end_color=self.colori['weekend'], fill_type='solid')

            if fill_giorno:
                 cell_data.fill = fill_giorno
                 # Applica sfondo a tutta la riga per chiarezza
                 for col_idx in range(2, len(nomi_addetti_ordinati) + 2):
                     ws.cell(riga, col_idx).fill = fill_giorno


            # Scrivi turni/stati per ogni addetto
            turni_del_giorno = calendario.get(giorno, {})
            for col, addetto in enumerate(nomi_addetti_ordinati, 2):
                cell = ws.cell(riga, col)
                cell.border = bordo_sottile
                cell.alignment = allineamento_centro # Centra il turno/stato

                stato_turno = turni_del_giorno.get(addetto, '-') # Default a '-' se manca l'addetto quel giorno

                fill_cella = None # Fill specifico per la cella (sovrascrive quello giorno)
                testo_cella = "-"
                font_cella = None

                if isinstance(stato_turno, (list, tuple)) and len(stato_turno) == 2:
                    # È un turno ('HH:MM', 'HH:MM')
                    testo_cella = f"{stato_turno[0]}-{stato_turno[1]}"
                    # Colora base a mattina/pomeriggio
                    try:
                         if self._get_orario_in_minuti(stato_turno[0]) < self._get_orario_in_minuti("13:00"):
                              fill_cella = PatternFill(start_color=self.colori['turno_mattina'], end_color=self.colori['turno_mattina'], fill_type='solid')
                         else:
                              fill_cella = PatternFill(start_color=self.colori['turno_pomeriggio'], end_color=self.colori['turno_pomeriggio'], fill_type='solid')
                    except: pass # Ignora errori di formato orario qui

                elif isinstance(stato_turno, str):
                    # È uno stato (FERIE, RIPOSO, FESTIVO, ERRORE...)
                    testo_cella = stato_turno
                    if stato_turno == 'FERIE':
                        fill_cella = PatternFill(start_color=self.colori['ferie'], end_color=self.colori['ferie'], fill_type='solid')
                    elif stato_turno == 'RIPOSO':
                        fill_cella = PatternFill(start_color=self.colori['riposo'], end_color=self.colori['riposo'], fill_type='solid')
                    elif stato_turno == 'FESTIVO':
                        # Usa lo stesso colore della riga festiva, ma assicurati sia applicato
                        fill_cella = PatternFill(start_color=self.colori['festivo'], end_color=self.colori['festivo'], fill_type='solid')
                    elif 'ERRORE' in stato_turno:
                         fill_cella = PatternFill(start_color=self.colori['errore'], end_color=self.colori['errore'], fill_type='solid')
                         font_cella = Font(color='FFFFFF', bold=True) # Testo bianco su sfondo rosso
                         testo_cella = "ERR!" # Testo corto per errore
                    else:
                         testo_cella = stato_turno # Mostra stringa com'è

                # Applica testo, fill e font
                cell.value = testo_cella
                if fill_cella: # Il fill specifico della cella ha la priorità
                    cell.fill = fill_cella
                elif fill_giorno: # Altrimenti usa il fill del giorno (weekend/festivo)
                     cell.fill = fill_giorno # Già applicato prima, ma riassicura coerenza
                     # Potrebbe essere necessario rimuovere il fill se è un turno normale in giorno festivo/weekend?
                     # if isinstance(stato_turno, (list, tuple)): cell.fill = PatternFill(fill_type=None) # Rimuovi fill giorno se c'è turno? Decisione stilistica.

                if font_cella:
                    cell.font = font_cella

        # Congela la prima riga (header)
        ws.freeze_panes = 'A2'

        # Salva il file
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        nome_file = os.path.join(desktop_path, f"Turni_{nome_mese}_{anno}.xlsx")
        try:
            wb.save(nome_file)
            messagebox.showinfo("Salvataggio Excel", f"File salvato con successo sul Desktop:\n{nome_file}")

            # Apri il file dopo salvataggio
            try:
                if os.name == 'nt': # Windows
                    os.startfile(nome_file)
                elif sys.platform == 'darwin': # macOS
                    subprocess.call(('open', nome_file))
                else: # linux variants
                    subprocess.call(('xdg-open', nome_file))
            except Exception as e_open:
                 print(f"Avviso: Impossibile aprire automaticamente il file Excel ({e_open})")
                 messagebox.showwarning("Apertura File", "File Excel salvato, ma impossibile aprirlo automaticamente.")

        except PermissionError:
             messagebox.showerror("Errore Salvataggio Excel", f"Permesso negato.\nIl file '{nome_file}' potrebbe essere aperto in un altro programma. Chiuderlo e riprovare.")
        except Exception as e_save:
            messagebox.showerror("Errore Salvataggio Excel", f"Errore durante il salvataggio del file Excel:\n{e_save}")
            print(f"Errore salvataggio Excel: {e_save}")
            traceback.print_exc()


    def visualizza_statistiche(self):
        """Visualizza le statistiche dei turni leggendo un file Excel generato"""
        # (Codice della funzione visualizza_statistiche originale)
        # ... (incolla qui il codice della funzione visualizza_statistiche dal file originale) ...
        # Cerca file Excel nella cartella Desktop (dove li salviamo)
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        try:
            files_turni = [f for f in os.listdir(desktop_path) if f.startswith('Turni_') and f.endswith('.xlsx')]
        except FileNotFoundError:
             messagebox.showerror("Errore", f"Cartella Desktop non trovata: {desktop_path}")
             return

        if not files_turni:
            messagebox.showinfo("Info", "Nessun file 'Turni_*.xlsx' trovato sul Desktop per generare statistiche.")
            return

        window = tk.Toplevel(self.root)
        window.title("Statistiche Turni da File Excel")
        window.geometry("800x600")

        # Frame per selezione file
        frame_select = ttk.Frame(window, padding=10)
        frame_select.pack(pady=10, fill=tk.X)

        ttk.Label(frame_select, text="Seleziona file Excel dal Desktop:").pack(side=tk.LEFT, padx=5)
        # Ordina i file per data (più recente prima)
        files_turni.sort(key=lambda name: os.path.getmtime(os.path.join(desktop_path, name)), reverse=True)
        file_var = tk.StringVar(value=files_turni[0] if files_turni else "")
        combo_files = ttk.Combobox(frame_select, textvariable=file_var, values=files_turni, state='readonly', width=40)
        combo_files.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Frame per mostrare le statistiche (con scrollbar)
        frame_stats_outer = ttk.Frame(window, padding=10)
        frame_stats_outer.pack(pady=10, fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(frame_stats_outer)
        scrollbar_stats = ttk.Scrollbar(frame_stats_outer, orient=tk.VERTICAL, command=canvas.yview)
        # Frame interno al canvas per contenere le etichette
        frame_stats_inner = ttk.Frame(canvas)

        # Configura scrollbar e canvas
        scrollbar_stats.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        canvas.configure(yscrollcommand=scrollbar_stats.set)
        canvas_window = canvas.create_window((0, 0), window=frame_stats_inner, anchor="nw")

        # Aggiorna scrollregion quando il frame interno cambia dimensione
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        frame_stats_inner.bind("<Configure>", on_frame_configure)

        # Aggiorna larghezza del frame interno quando il canvas cambia dimensione
        def on_canvas_configure(event):
             canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", on_canvas_configure)

        # Funzione per attivare scroll con rotellina mouse
        def on_mouse_wheel(event):
             # Su Windows/macOS delta è +/- 120, su Linux è +/- 1 (event.num)
             if event.num == 4 or event.delta > 0: # Linux scroll up or Win/Mac scroll up
                 canvas.yview_scroll(-1, "units")
             elif event.num == 5 or event.delta < 0: # Linux scroll down or Win/Mac scroll down
                 canvas.yview_scroll(1, "units")

        # Binding per diverse piattaforme
        canvas.bind_all("<MouseWheel>", on_mouse_wheel) # Windows
        canvas.bind_all("<Button-4>", on_mouse_wheel) # Linux scroll up
        canvas.bind_all("<Button-5>", on_mouse_wheel) # Linux scroll down


        def aggiorna_statistiche():
            """Aggiorna le statistiche visualizzate leggendo l'Excel selezionato."""
            # Pulisci frame statistiche precedente
            for widget in frame_stats_inner.winfo_children():
                widget.destroy()

            file_selezionato = file_var.get()
            if not file_selezionato:
                ttk.Label(frame_stats_inner, text="Selezionare un file.").pack()
                return

            file_path = os.path.join(desktop_path, file_selezionato)

            try:
                    # ---> INIZIO CODICE AGGIUNTO/MODIFICATO <---
                # Estrai anno e mese dal nome del file (es. Turni_Maggio_2025.xlsx)
                try:
                    # Prova a separare nome, mese, anno
                    parts = file_selezionato.replace('.xlsx','').split('_')
                    if len(parts) >= 3:
                        anno_stats = int(parts[-1])
                        # Converti nome mese italiano in numero mese
                        nome_mese_stats = parts[-2].capitalize()
                        # Assicurati che mesi_italiano sia definito come in genera_pianificazione
                        mesi_italiano = [""] + list(calendar.month_name)[1:]
                        try:
                            import locale
                            # Tentativo impostazione locale per nomi mese
                            try:
                                    locale.setlocale(locale.LC_TIME, 'it_IT.UTF-8')
                            except locale.Error: # Fallback se locale non supportato
                                try: # try:
                                        locale.setlocale(locale.LC_TIME, 'Italian_Italy') # Windows?
                                except:
                                        pass # Usa inglese se fallisce
                            mesi_italiano = [""] + [calendar.month_name[i].capitalize() for i in range(1, 13)]
                        except ImportError:
                            print("Modulo locale non trovato, uso nomi mese inglesi.")
                            pass

                        try:
                            mese_stats = mesi_italiano.index(nome_mese_stats)
                            if mese_stats == 0: raise ValueError
                        except ValueError:
                            raise ValueError(f"Nome mese '{nome_mese_stats}' non riconosciuto.")

                        # Ottieni i festivi per quel mese/anno
                        festivi_mese_corrente = self._get_festivi_mese(anno_stats, mese_stats)
                        print(f"Statistiche: Calcolati festivi per {nome_mese_stats} {anno_stats}: {festivi_mese_corrente}")
                    else:
                        raise ValueError("Formato nome file non riconosciuto (atteso Turni_Mese_Anno.xlsx)")
                except Exception as e_parse:
                    messagebox.showerror("Errore Nome File", f"Impossibile determinare mese/anno dal nome file '{file_selezionato}'.\nErrore: {e_parse}\nImpossibile calcolare festivi lavorati.")
                    # Definisci comunque la variabile come set vuoto per evitare il NameError
                    festivi_mese_corrente = set()
                # ---> FINE CODICE AGGIUNTO/MODIFICATO <---
                # Carica dati Excel usando pandas
                df = pd.read_excel(file_path, index_col=None) # Legge la prima riga come header

                # Verifica colonne attese
                if not df.columns.to_list() or not df.columns[0].lower() == 'data':
                     ttk.Label(frame_stats_inner, text=f"Errore: Il file '{file_selezionato}' non sembra avere il formato atteso (manca colonna 'Data'?).").pack(padx=5, pady=5)
                     return

                nomi_addetti_excel = df.columns[1:] # Nomi addetti dalle colonne Excel

                ttk.Label(frame_stats_inner, text=f"Statistiche per: {file_selezionato}", font=('Helvetica', 12, 'bold')).pack(pady=10)

                # Calcola statistiche per ogni addetto presente nel file Excel
                for addetto in nomi_addetti_excel:
                    # Inizializza contatori per questo addetto
                    turni_totali = 0
                    ferie = 0
                    riposi = 0
                    festivi_lavorati = 0 # Contiamo turni nei giorni festivi
                    ore_totali = 0.0
                    domeniche_lavorate = 0
                    errori_cella = 0
                    giorni_effettivi_lavorati = 0 # Conta giorni con un turno valido

                    # Analizziamo ogni riga (giorno) per questo addetto
                    for index, riga in df.iterrows():
                        # Ottieni la data dalla prima colonna
                        try:
                            # Prova diversi formati comuni per la data (es. 'gg/mm/aaaa (Gio)')
                            data_str = str(riga.iloc[0]).split('(')[0].strip()
                            data = datetime.strptime(data_str, '%d/%m/%Y')
                        except (ValueError, TypeError):
                            # print(f"Riga {index+2}: Impossibile interpretare la data '{riga.iloc[0]}'. Salto riga.")
                            continue # Salta riga se la data non è interpretabile

                        # Determina se è festivo (usando la stessa logica della generazione)
                        data_str_dm = data.strftime('%d-%m')
                        is_festivo = data_str_dm in festivi_mese_corrente

                        # Ottieni il valore della cella per l'addetto corrente
                        try:
                             valore_cella = str(riga[addetto]).strip() if pd.notna(riga[addetto]) else "-"
                        except KeyError:
                             # print(f"Addetto '{addetto}' non trovato nel file Excel alla riga {index+2}")
                             continue # Addetto non presente nel file?

                        # Analizza il contenuto della cella
                        if valore_cella == "FERIE":
                            ferie += 1
                        elif valore_cella == "RIPOSO":
                            riposi += 1
                        elif valore_cella == "FESTIVO":
                             # Non conta come giorno lavorato, ma potrebbe essere utile saperlo
                             pass
                        elif valore_cella == "-" or valore_cella == "":
                             # Giorno non lavorato (non ferie/riposo/festivo)
                             pass
                        elif 'ERR' in valore_cella: # Cattura celle marcate come Errore
                            errori_cella += 1
                        elif '-' in valore_cella and ':' in valore_cella:
                            # Probabilmente è un turno "HH:MM-HH:MM"
                            try:
                                inizio_str, fine_str = valore_cella.split('-')
                                inizio_min = self._get_orario_in_minuti(inizio_str.strip())
                                fine_min = self._get_orario_in_minuti(fine_str.strip())

                                if inizio_min is not None and fine_min is not None:
                                    turni_totali += 1
                                    giorni_effettivi_lavorati += 1
                                    durata_min = fine_min - inizio_min
                                    if durata_min < 0: durata_min += 24 * 60 # Mezzanotte
                                    ore_totali += durata_min / 60.0

                                    # Controlla se è domenica
                                    if data.weekday() == 6: # Domenica = 6
                                        domeniche_lavorate += 1
                                    # Controlla se è festivo lavorato
                                    if is_festivo:
                                         festivi_lavorati += 1
                                else:
                                     # Formato ora non valido dentro il turno?
                                     # print(f"Addetto {addetto}, Giorno {data_str}: formato ora non valido in '{valore_cella}'")
                                     errori_cella += 1
                            except ValueError:
                                # Errore nello split o formato?
                                # print(f"Addetto {addetto}, Giorno {data_str}: impossibile analizzare turno '{valore_cella}'")
                                errori_cella += 1
                        else:
                             # Valore non riconosciuto
                             # print(f"Addetto {addetto}, Giorno {data_str}: valore cella non riconosciuto '{valore_cella}'")
                             errori_cella += 1


                    # Crea frame per le statistiche dell'addetto
                    frame_addetto = ttk.LabelFrame(frame_stats_inner, text=addetto, padding=10)
                    frame_addetto.pack(fill=tk.X, padx=10, pady=5)

                    # Mostra statistiche calcolate
                    ttk.Label(frame_addetto, text=f"Giorni Lavorati: {giorni_effettivi_lavorati}").grid(row=0, column=0, sticky='w', padx=5)
                    ttk.Label(frame_addetto, text=f"Turni Totali: {turni_totali}").grid(row=0, column=1, sticky='w', padx=5)
                    ttk.Label(frame_addetto, text=f"Ore Lavorate: {ore_totali:.2f}").grid(row=0, column=2, sticky='w', padx=5)

                    ttk.Label(frame_addetto, text=f"Giorni Ferie: {ferie}").grid(row=1, column=0, sticky='w', padx=5)
                    ttk.Label(frame_addetto, text=f"Giorni Riposo: {riposi}").grid(row=1, column=1, sticky='w', padx=5)

                    ttk.Label(frame_addetto, text=f"Domeniche Lavorate: {domeniche_lavorate}").grid(row=2, column=0, sticky='w', padx=5)
                    ttk.Label(frame_addetto, text=f"Festivi Lavorati: {festivi_lavorati}").grid(row=2, column=1, sticky='w', padx=5)

                    if errori_cella > 0:
                         ttk.Label(frame_addetto, text=f"Errori/Valori Sconosciuti: {errori_cella}", foreground='red').grid(row=2, column=2, sticky='w', padx=5)


                    # Verifica rispetto monte ore (se l'addetto è nei dati correnti dell'app)
                    if addetto in self.addetti:
                        info_addetto = self.addetti[addetto]
                        ore_contratto = info_addetto.get('ore_contratto', 0)
                        ore_max = info_addetto.get('ore_max', 0)
                        straordinario_ok = info_addetto.get('straordinario', False)
                        stato_ore = ""
                        colore_stato = "black"

                        # Calcola media ore settimanali (approssimata)
                        num_settimane = df.shape[0] / 7.0 # Numero giorni / 7
                        media_ore_sett = ore_totali / num_settimane if num_settimane > 0 else 0

                        # Verifica limite massimo settimanale (più significativo del totale mensile)
                        if ore_max > 0 and media_ore_sett > (ore_max + 0.1) and not straordinario_ok: # Tolleranza 0.1 ore
                             stato_ore = f"⚠️ ATTENZIONE: Media ore sett. ({media_ore_sett:.1f}) > Max ({ore_max}) senza autorizzazione straordinari!"
                             colore_stato = "red"
                        elif ore_max > 0 and media_ore_sett > (ore_max + 0.1):
                             stato_ore = f"INFO: Media ore sett. ({media_ore_sett:.1f}) > Max ({ore_max}), con autorizzazione straordinari."
                             colore_stato = "darkorange"
                        # elif ore_contratto > 0 and media_ore_sett < (ore_contratto - 0.1):
                        #      stato_ore = f"INFO: Media ore sett. ({media_ore_sett:.1f}) < Contratto ({ore_contratto})."
                        #      colore_stato = "blue"

                        if stato_ore:
                             ttk.Label(frame_addetto, text=stato_ore, foreground=colore_stato).grid(row=3, column=0, columnspan=3, sticky='w', padx=5, pady=5)
                    else:
                         ttk.Label(frame_addetto, text="INFO: Addetto non trovato nei dati correnti dell'applicazione per verifica ore.", foreground='gray').grid(row=3, column=0, columnspan=3, sticky='w', padx=5, pady=5)

                    # Configura colonne del frame addetto per allineamento
                    frame_addetto.grid_columnconfigure(0, weight=1)
                    frame_addetto.grid_columnconfigure(1, weight=1)
                    frame_addetto.grid_columnconfigure(2, weight=1)


            except FileNotFoundError:
                 ttk.Label(frame_stats_inner, text=f"Errore: File non trovato '{file_path}'.").pack(padx=5, pady=5)
            except Exception as e:
                 ttk.Label(frame_stats_inner, text=f"Errore imprevisto nell'analisi del file:\n{e}").pack(padx=5, pady=5)
                 print(f"Errore analisi statistiche: {e}")
                 traceback.print_exc()

            # Aggiorna la scrollregion dopo aver aggiunto tutto
            canvas.update_idletasks() # Assicura che le dimensioni siano aggiornate
            canvas.config(scrollregion=canvas.bbox("all"))

        # Bottone per aggiornare statistiche
        btn_aggiorna = ttk.Button(frame_select, text="Mostra Statistiche", command=aggiorna_statistiche)
        btn_aggiorna.pack(side=tk.LEFT, padx=10)

        # Aggiorna statistiche iniziali se un file è preselezionato
        if file_var.get():
            aggiorna_statistiche()


    def genera_pianificazione(self):
        """Genera la pianificazione dei turni per il mese selezionato"""
        if not self.addetti:
             messagebox.showerror("Errore", "Nessun addetto definito. Aggiungere addetti prima di generare.")
             return
        if not self.turni_disponibili:
             messagebox.showerror("Errore", "Nessun turno definito. Aggiungere turni disponibili prima di generare.")
             return


        window = tk.Toplevel(self.root)
        window.title("Genera Pianificazione Mensile")
        window.geometry("400x200") # Ridotta finestra

        # Frame per selezione periodo
        frame_periodo = ttk.LabelFrame(window, text="Seleziona Periodo", padding=10)
        frame_periodo.pack(pady=10, padx=10, fill='x')

        # Anno
        ttk.Label(frame_periodo, text="Anno:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        current_year = datetime.now().year
        anno_var = tk.IntVar(value=current_year)
        # Permette anni ragionevoli
        ttk.Spinbox(frame_periodo, from_=current_year - 5, to=current_year + 5, textvariable=anno_var, width=6).grid(row=0, column=1, padx=5, pady=5, sticky='w')

        # Mese
        ttk.Label(frame_periodo, text="Mese:").grid(row=0, column=2, padx=5, pady=5, sticky='w')
        mesi_italiano = [""] + list(calendar.month_name)[1:] # Usa nomi mese completi
        try:
            # Imposta locale italiano per nomi mese, se possibile
            import locale
            locale.setlocale(locale.LC_TIME, 'it_IT.UTF-8')
            mesi_italiano = [""] + [calendar.month_name[i].capitalize() for i in range(1, 13)]
        except:
            print("Locale italiano non disponibile, usando nomi mese default.")
            pass # Usa nomi default se locale non trovato

        mese_var = tk.StringVar(value=mesi_italiano[datetime.now().month])
        ttk.Combobox(frame_periodo, textvariable=mese_var,
                    values=mesi_italiano[1:], state='readonly', width=12).grid(row=0, column=3, padx=5, pady=5, sticky='w')

        # Funzione interna chiamata dal bottone
        def genera():
            """Genera i turni per il mese selezionato"""
            try:
                anno = anno_var.get()
                # Trova l'indice del mese selezionato (1-12)
                mese_nome_selezionato = mese_var.get()
                try:
                     mese = mesi_italiano.index(mese_nome_selezionato)
                     if mese == 0: raise ValueError # Indice 0 non è un mese valido
                except ValueError:
                     messagebox.showerror("Errore Interno", "Mese selezionato non valido.")
                     return

                print(f"Avvio generazione per {mese_nome_selezionato} {anno}...")
                # Disabilita bottone durante generazione? (Opzionale)
                btn_genera.config(state='disabled', text='Generazione in corso...')
                window.update_idletasks() # Forza aggiornamento UI

                # ----> CHIAMA LA NUOVA FUNZIONE DI PIANIFICAZIONE <----
                calendario = self._genera_calendario_mensile_refactored(anno, mese)

                print("Generazione calendario completata.")
                # Riabilita bottone
                btn_genera.config(state='normal', text='Genera Pianificazione')


                # Controlla se il calendario contiene errori critici (es. copertura incompleta)
                contiene_errori = False
                for giorno, dati_giorno in calendario.items():
                    if isinstance(dati_giorno, dict):
                        if any('ERRORE' in str(v) for v in dati_giorno.values()):
                             contiene_errori = True
                             break
                if contiene_errori:
                    print("ATTENZIONE: La pianificazione contiene errori o coperture incomplete.")
                    if not messagebox.askyesno("Attenzione", "La pianificazione generata contiene errori o coperture incomplete (verificare log e file Excel).\n\nSalvare comunque il file Excel?", icon='warning'):
                        print("Salvataggio annullato dall'utente.")
                        window.destroy() # Chiudi finestra generazione
                        return # Non salvare se l'utente dice no

                print("Avvio salvataggio Excel...")
                # Salva su Excel (la funzione ora gestisce apertura e messaggi)
                self._salva_calendario_excel(calendario, anno, mese)

                print("Salvataggio Excel completato.")
                # Messaggio finale spostato dentro _salva_calendario_excel
                window.destroy() # Chiudi la finestra di generazione

            except ValueError:
                 messagebox.showerror("Errore Input", "Anno o Mese non valido.")
                 if 'btn_genera' in locals(): btn_genera.config(state='normal', text='Genera Pianificazione')
            except Exception as e:
                 messagebox.showerror("Errore Inaspettato", f"Si è verificato un errore imprevisto durante la generazione:\n{str(e)}")
                 print("--- ERRORE INASPETTATO ---")
                 traceback.print_exc() # Stampa l'errore completo nella console per debug
                 if 'btn_genera' in locals(): btn_genera.config(state='normal', text='Genera Pianificazione')


        # Bottone per avviare la generazione
        btn_genera = ttk.Button(window, text="Genera Pianificazione", command=genera)
        btn_genera.pack(pady=20)


    def run(self):
        """Avvia l'applicazione Tkinter"""
        self.root.mainloop()

# ==========================================================================
# Avvio dell'applicazione
# ==========================================================================
if __name__ == "__main__":
    # Import aggiuntivi per l'apertura file cross-platform in _salva_calendario_excel
    import sys
    import subprocess

    app = GestioneTurni()
    app.run()
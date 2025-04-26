import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime, timedelta
import calendar
import random
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

class GestioneTurni:
    def __init__(self):
        """Inizializzazione dell'applicazione"""
        # Inizializzazione delle variabili principali
        self.addetti = {}  # Dizionario per memorizzare i dati degli addetti
        self.turni_disponibili = []  # Lista dei turni disponibili
        self.giorni_festivi = [
            "01-01",  # Capodanno
            "20-04",  # 20 aprile
            "01-05",  # 1 maggio
            "25-12",  # Natale
            "26-12"   # Santo Stefano
        ]
        
        # Orari di apertura del supermercato
        self.orario_apertura = "08:00"
        self.orario_chiusura = "21:00"
        
        # Colori per Excel
        self.colori = {
            'header': 'CCE5FF',     # Azzurro chiaro per header
            'weekend': 'FFE6E6',    # Rosa chiaro per weekend
            'turno_mattina': 'E6FFE6',  # Verde chiaro per turni mattina
            'turno_pomeriggio': 'FFE6CC',  # Arancione chiaro per turni pomeriggio
            'riposo': 'F2F2F2',     # Grigio chiaro per riposi
            'ferie': 'FFFF99',      # Giallo chiaro per ferie
            'festivo': 'FF9999'     # Rosso chiaro per festivi
        }
        
        # Carica i dati se esistono
        self.carica_dati()
        
        # Creazione della finestra principale
        self.root = tk.Tk()
        self.root.title("Gestione Turni Supermercato")
        self.root.geometry("800x600")
        
        # Creazione del menu principale
        self.crea_menu_principale()
    
    def carica_dati(self):
        """Carica i dati salvati se esistono"""
        try:
            if os.path.exists('dati_turni.json'):
                with open('dati_turni.json', 'r') as f:
                    dati = json.load(f)
                    self.addetti = dati.get('addetti', {})
                    self.turni_disponibili = dati.get('turni', [])
        except Exception as e:
            print(f"Errore nel caricamento dei dati: {e}")
    
    def salva_dati(self):
        """Salva i dati su file"""
        try:
            dati = {
                'addetti': self.addetti,
                'turni': self.turni_disponibili
            }
            with open('dati_turni.json', 'w') as f:
                json.dump(dati, f)
        except Exception as e:
            print(f"Errore nel salvataggio dei dati: {e}")
    
    def crea_menu_principale(self):
        """Crea il menu principale dell'applicazione"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        ttk.Button(main_frame, text="Gestione Addetti", 
                  command=self.gestione_addetti).grid(row=0, column=0, pady=5)
        ttk.Button(main_frame, text="Gestione Turni", 
                  command=self.gestione_turni).grid(row=1, column=0, pady=5)
        ttk.Button(main_frame, text="Gestione Ferie e Riposi", 
                  command=self.gestione_ferie_riposi).grid(row=2, column=0, pady=5)
        ttk.Button(main_frame, text="Genera Pianificazione", 
                  command=self.genera_pianificazione).grid(row=3, column=0, pady=5)
        ttk.Button(main_frame, text="Visualizza Statistiche", 
                  command=self.visualizza_statistiche).grid(row=4, column=0, pady=5)
    
    def gestione_addetti(self):
        """Gestisce l'aggiunta e la modifica degli addetti"""
        window = tk.Toplevel(self.root)
        window.title("Gestione Addetti")
        window.geometry("600x500")
        
        # Frame per la lista degli addetti esistenti
        frame_lista = ttk.Frame(window)
        frame_lista.grid(row=0, column=0, padx=10, pady=10)
        
        # Lista degli addetti esistenti
        ttk.Label(frame_lista, text="Addetti esistenti:").grid(row=0, column=0)
        lista_addetti = tk.Listbox(frame_lista, width=30, height=10)
        lista_addetti.grid(row=1, column=0)
        for addetto in self.addetti:
            lista_addetti.insert(tk.END, addetto)
        
        # Frame per il form di inserimento
        frame_form = ttk.Frame(window)
        frame_form.grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(frame_form, text="Nome:").grid(row=0, column=0, pady=5)
        nome_var = tk.StringVar()
        ttk.Entry(frame_form, textvariable=nome_var).grid(row=0, column=1, pady=5)
        
        ttk.Label(frame_form, text="Ore Contratto:").grid(row=1, column=0, pady=5)
        ore_var = tk.IntVar(value=40)
        ttk.Entry(frame_form, textvariable=ore_var).grid(row=1, column=1, pady=5)
        
        ttk.Label(frame_form, text="Ore Max:").grid(row=2, column=0, pady=5)
        ore_max_var = tk.IntVar(value=48)
        ttk.Entry(frame_form, textvariable=ore_max_var).grid(row=2, column=1, pady=5)
        
        ttk.Label(frame_form, text="Straordinario:").grid(row=3, column=0, pady=5)
        straordinario_var = tk.BooleanVar()
        ttk.Checkbutton(frame_form, variable=straordinario_var).grid(row=3, column=1, pady=5)
        
        def salva_addetto():
            nome = nome_var.get()
            if nome:
                self.addetti[nome] = {
                    'ore_contratto': ore_var.get(),
                    'ore_max': ore_max_var.get(),
                    'straordinario': straordinario_var.get(),
                    'giorni_riposo': [],
                    'ferie': []
                }
                lista_addetti.insert(tk.END, nome)
                self.salva_dati()
                messagebox.showinfo("Successo", f"Addetto {nome} aggiunto correttamente")
                
        def elimina_addetto():
            selection = lista_addetti.curselection()
            if selection:
                nome = lista_addetti.get(selection[0])
                if messagebox.askyesno("Conferma", f"Vuoi eliminare l'addetto {nome}?"):
                    del self.addetti[nome]
                    lista_addetti.delete(selection[0])
                    self.salva_dati()
        
        ttk.Button(frame_form, text="Salva", 
                  command=salva_addetto).grid(row=4, column=0, columnspan=2, pady=10)
        ttk.Button(frame_form, text="Elimina Selezionato", 
                  command=elimina_addetto).grid(row=5, column=0, columnspan=2, pady=10)
    
    def gestione_turni(self):
        """Gestisce la definizione dei turni disponibili"""
        window = tk.Toplevel(self.root)
        window.title("Gestione Turni")
        window.geometry("500x400")
        
        # Lista dei turni esistenti
        frame_lista = ttk.Frame(window)
        frame_lista.grid(row=0, column=0, padx=10, pady=10)
        
        ttk.Label(frame_lista, text="Turni esistenti:").grid(row=0, column=0)
        lista_turni = tk.Listbox(frame_lista, width=30, height=10)
        lista_turni.grid(row=1, column=0)
        for turno in self.turni_disponibili:
            lista_turni.insert(tk.END, f"{turno[0]} - {turno[1]}")
        
        # Form per nuovo turno
        frame_form = ttk.Frame(window)
        frame_form.grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(frame_form, text="Ora Inizio (HH:MM):").grid(row=0, column=0, pady=5)
        inizio_var = tk.StringVar()
        ttk.Entry(frame_form, textvariable=inizio_var).grid(row=0, column=1, pady=5)
        
        ttk.Label(frame_form, text="Ora Fine (HH:MM):").grid(row=1, column=0, pady=5)
        fine_var = tk.StringVar()
        ttk.Entry(frame_form, textvariable=fine_var).grid(row=1, column=1, pady=5)
        
        def valida_orario(orario):
            """Valida il formato dell'orario"""
            try:
                ore, minuti = map(int, orario.split(':'))
                return 0 <= ore <= 23 and 0 <= minuti <= 59
            except:
                return False
        
        def salva_turno():
            inizio = inizio_var.get()
            fine = fine_var.get()
            if not (valida_orario(inizio) and valida_orario(fine)):
                messagebox.showerror("Errore", "Formato orario non valido (HH:MM)")
                return
            if inizio >= fine:
                messagebox.showerror("Errore", "L'ora di inizio deve essere precedente all'ora di fine")
                return
            if inizio < self.orario_apertura or fine > self.orario_chiusura:
                messagebox.showerror("Errore", f"Il turno deve essere tra {self.orario_apertura} e {self.orario_chiusura}")
                return
            
            self.turni_disponibili.append((inizio, fine))
            lista_turni.insert(tk.END, f"{inizio} - {fine}")
            self.salva_dati()
            messagebox.showinfo("Successo", f"Turno {inizio}-{fine} aggiunto")
        
        def elimina_turno():
            selection = lista_turni.curselection()
            if selection:
                idx = selection[0]
                if messagebox.askyesno("Conferma", "Vuoi eliminare il turno selezionato?"):
                    del self.turni_disponibili[idx]
                    lista_turni.delete(idx)
                    self.salva_dati()
        
        ttk.Button(frame_form, text="Aggiungi Turno", 
                  command=salva_turno).grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(frame_form, text="Elimina Selezionato", 
                  command=elimina_turno).grid(row=3, column=0, columnspan=2, pady=10)

    def gestione_ferie_riposi(self):
        """Gestisce ferie e giorni di riposo degli addetti"""
        if not self.addetti:
            messagebox.showerror("Errore", "Inserire prima gli addetti")
            return
        
        window = tk.Toplevel(self.root)
        window.title("Gestione Ferie e Riposi")
        window.geometry("800x600")
        
        # Frame principale diviso in due colonne
        frame_sx = ttk.Frame(window)
        frame_dx = ttk.Frame(window)
        frame_sx.grid(row=0, column=0, padx=10, pady=10, sticky='nsew')
        frame_dx.grid(row=0, column=1, padx=10, pady=10, sticky='nsew')
        
        # Configurazione del grid
        window.grid_columnconfigure(0, weight=1)
        window.grid_columnconfigure(1, weight=1)
        
        # ---- Sezione Selezione Addetto (comune) ----
        frame_select = ttk.LabelFrame(frame_sx, text="Selezione Addetto")
        frame_select.pack(fill='x', pady=5)
        
        ttk.Label(frame_select, text="Seleziona Addetto:").pack(side=tk.LEFT, padx=5)
        addetto_var = tk.StringVar()
        combo_addetti = ttk.Combobox(frame_select, textvariable=addetto_var, 
                                   values=list(self.addetti.keys()))
        combo_addetti.pack(side=tk.LEFT, padx=5)
        
        # ---- Sezione Gestione Ferie (sinistra) ----
        frame_ferie = ttk.LabelFrame(frame_sx, text="Gestione Ferie")
        frame_ferie.pack(fill='both', expand=True, pady=5)
        
        # Calendario per selezione ferie
        ttk.Label(frame_ferie, text="Seleziona Data:").pack(pady=5)
        
        frame_calendario = ttk.Frame(frame_ferie)
        frame_calendario.pack(pady=5)
        
        # Selezione anno e mese
        anno_var = tk.StringVar(value=str(datetime.now().year))
        mese_var = tk.StringVar(value=str(datetime.now().month))
        
        ttk.Label(frame_calendario, text="Anno:").grid(row=0, column=0, padx=5)
        ttk.Entry(frame_calendario, textvariable=anno_var, width=6).grid(row=0, column=1)
        ttk.Label(frame_calendario, text="Mese:").grid(row=0, column=2, padx=5)
        ttk.Spinbox(frame_calendario, from_=1, to=12, width=4, 
                   textvariable=mese_var).grid(row=0, column=3)
        
        # Calendario per selezione giorni
        frame_giorni = ttk.Frame(frame_ferie)
        frame_giorni.pack(pady=5)
        giorni_vars = {}  # Variabili per i checkbutton dei giorni
        
        def aggiorna_calendario():
            """Aggiorna la visualizzazione del calendario"""
            # Pulisci frame giorni
            for widget in frame_giorni.winfo_children():
                widget.destroy()
            
            try:
                anno = int(anno_var.get())
                mese = int(mese_var.get())
                
                # Crea intestazioni giorni settimana
                giorni_settimana = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
                for i, giorno in enumerate(giorni_settimana):
                    ttk.Label(frame_giorni, text=giorno).grid(row=0, column=i, padx=2)
                
                # Ottieni il calendario del mese
                cal = calendar.monthcalendar(anno, mese)
                giorni_vars.clear()
                
                # Crea checkbutton per ogni giorno
                for settimana in range(len(cal)):
                    for giorno in range(7):
                        day_num = cal[settimana][giorno]
                        if day_num != 0:
                            var = tk.BooleanVar()
                            giorni_vars[day_num] = var
                            cb = ttk.Checkbutton(frame_giorni, text=str(day_num),
                                               variable=var)
                            cb.grid(row=settimana+1, column=giorno, padx=2)
                
                # Evidenzia le ferie già programmate
                addetto = addetto_var.get()
                if addetto in self.addetti:
                    for data in self.addetti[addetto]['ferie']:
                        data_obj = datetime.strptime(data, '%Y-%m-%d')
                        if data_obj.year == anno and data_obj.month == mese:
                            if data_obj.day in giorni_vars:
                                giorni_vars[data_obj.day].set(True)
                
            except ValueError:
                messagebox.showerror("Errore", "Data non valida")
        
        ttk.Button(frame_calendario, text="Aggiorna Calendario", 
                  command=aggiorna_calendario).grid(row=0, column=4, padx=5)
        
        # Lista delle ferie già programmate
        ttk.Label(frame_ferie, text="Ferie Programmate:").pack(pady=5)
        lista_ferie = tk.Listbox(frame_ferie, height=6)
        lista_ferie.pack(fill='x', padx=5)
        
        def salva_ferie():
            """Salva le ferie selezionate per l'addetto"""
            addetto = addetto_var.get()
            if not addetto:
                messagebox.showerror("Errore", "Selezionare un addetto")
                return
                
            try:
                anno = int(anno_var.get())
                mese = int(mese_var.get())
                
                # Raccogli i giorni selezionati
                nuove_ferie = []
                for giorno, var in giorni_vars.items():
                    if var.get():
                        data = datetime(anno, mese, giorno)
                        nuove_ferie.append(data.strftime('%Y-%m-%d'))
                
                # Aggiorna le ferie dell'addetto
                self.addetti[addetto]['ferie'] = nuove_ferie
                self.salva_dati()
                aggiorna_liste()
                messagebox.showinfo("Successo", "Ferie salvate correttamente")
                
            except ValueError:
                messagebox.showerror("Errore", "Data non valida")
        
        ttk.Button(frame_ferie, text="Salva Ferie Selezionate", 
                  command=salva_ferie).pack(pady=5)
        
        # ---- Sezione Gestione Riposi (destra) ----
        frame_riposi = ttk.LabelFrame(frame_dx, text="Gestione Giorni di Riposo")
        frame_riposi.pack(fill='both', expand=True, pady=5)
        
        # Checkbox per ogni giorno della settimana
        giorni_settimana_completi = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", 
                                   "Venerdì", "Sabato", "Domenica"]
        riposi_vars = []
        
        for i, giorno in enumerate(giorni_settimana_completi):
            var = tk.BooleanVar()
            riposi_vars.append(var)
            ttk.Checkbutton(frame_riposi, text=giorno, 
                          variable=var).pack(anchor='w', padx=10)
        
        # Lista dei riposi programmati
        ttk.Label(frame_riposi, text="Giorni di Riposo Attuali:").pack(pady=5)
        lista_riposi = tk.Listbox(frame_riposi, height=6)
        lista_riposi.pack(fill='x', padx=5)
        
        def salva_riposi():
            """Salva i giorni di riposo per l'addetto"""
            addetto = addetto_var.get()
            if not addetto:
                messagebox.showerror("Errore", "Selezionare un addetto")
                return
            
            giorni_riposo = [i for i, var in enumerate(riposi_vars) if var.get()]
            if giorni_riposo:
                self.addetti[addetto]['giorni_riposo'] = giorni_riposo
                self.salva_dati()
                aggiorna_liste()
                messagebox.showinfo("Successo", "Giorni di riposo salvati")
            else:
                messagebox.showerror("Errore", "Selezionare almeno un giorno di riposo")
        
        ttk.Button(frame_riposi, text="Salva Giorni di Riposo", 
                  command=salva_riposi).pack(pady=5)
        
        def aggiorna_liste():
            """Aggiorna le liste di ferie e riposi"""
            addetto = addetto_var.get()
            if addetto in self.addetti:
                # Aggiorna lista ferie
                lista_ferie.delete(0, tk.END)
                for data in sorted(self.addetti[addetto]['ferie']):
                    lista_ferie.insert(tk.END, data)
                
                # Aggiorna lista riposi
                lista_riposi.delete(0, tk.END)
                for giorno in sorted(self.addetti[addetto]['giorni_riposo']):
                    lista_riposi.insert(tk.END, giorni_settimana_completi[giorno])
        
        def on_select_addetto(event):
            """Callback quando viene selezionato un addetto"""
            aggiorna_liste()
            aggiorna_calendario()
            
            # Aggiorna checkbox riposi
            addetto = addetto_var.get()
            if addetto in self.addetti:
                for i, var in enumerate(riposi_vars):
                    var.set(i in self.addetti[addetto]['giorni_riposo'])
        
        # Binding eventi
        combo_addetti.bind('<<ComboboxSelected>>', on_select_addetto)
        
        # Inizializza il calendario
        aggiorna_calendario()

    def _controlla_copertura_oraria(self, turni_giorno):
        """
        Controlla la copertura oraria per un giorno dato l'insieme dei turni assegnati.
        Restituisce una lista di intervalli temporali non coperti.
        """
        def orario_in_minuti(orario):
            ore, minuti = map(int, orario.split(':'))
            return ore * 60 + minuti
        
        apertura = orario_in_minuti(self.orario_apertura)
        chiusura = orario_in_minuti(self.orario_chiusura)
        
        # Crea un array di minuti per tracciare la copertura
        copertura = [False] * (chiusura - apertura)
        
        # Segna i minuti coperti dai turni assegnati
        for turno in turni_giorno.values():
            inizio = orario_in_minuti(turno[0]) - apertura
            fine = orario_in_minuti(turno[1]) - apertura
            
            # Assicuriamoci che gli indici siano validi
            inizio = max(0, inizio)
            fine = min(len(copertura), fine)
            
            for i in range(inizio, fine):
                copertura[i] = True
        
        # Trova gli intervalli non coperti
        buchi = []
        inizio_buco = None
        
        for i, coperto in enumerate(copertura):
            if not coperto and inizio_buco is None:
                inizio_buco = i
            elif coperto and inizio_buco is not None:
                buchi.append((inizio_buco + apertura, i + apertura))
                inizio_buco = None
                
        if inizio_buco is not None:
            buchi.append((inizio_buco + apertura, len(copertura) + apertura))
        
        return buchi

    def _calcola_punteggio_turno(self, addetto, turno, giorno, mese, anno, turni_assegnati):
        """Calcola un punteggio per l'assegnazione di un turno a un addetto."""
        punteggio = 0
        
        # Calcola ore già lavorate
        ore_lavorate = 0
        for g, turni in turni_assegnati.items():
            if addetto in turni:
                inizio = datetime.strptime(turni[addetto][0], '%H:%M')
                fine = datetime.strptime(turni[addetto][1], '%H:%M')
                ore_lavorate += (fine - inizio).seconds / 3600
        
        # Gestione ore contrattuali
        ore_contratto = self.addetti[addetto]['ore_contratto']
        ore_max = self.addetti[addetto]['ore_max']
        
        # Calcola ore del turno attuale
        inizio_turno = datetime.strptime(turno[0], '%H:%M')
        fine_turno = datetime.strptime(turno[1], '%H:%M')
        ore_turno = (fine_turno - inizio_turno).seconds / 3600
        
        # Bonus se le ore sono sotto il contratto
        if ore_lavorate + ore_turno <= ore_contratto:
            punteggio += 10
        
        # Malus se si supera il massimo di ore (se non è permesso lo straordinario)
        if not self.addetti[addetto]['straordinario'] and ore_lavorate + ore_turno > ore_max:
            punteggio -= 20
        
        # Controlla se l'addetto ha fatto lo stesso turno nei giorni precedenti
        turno_str = f"{turno[0]}-{turno[1]}"
        turni_precedenti = []
        
        # Controlla gli ultimi 5 giorni (aumentato da 3 a 5 per evitare ripetizioni)
        for g in range(max(1, giorno-5), giorno):
            if g in turni_assegnati and addetto in turni_assegnati[g]:
                t = turni_assegnati[g][addetto]
                turni_precedenti.append(f"{t[0]}-{t[1]}")
        
        # Malus MOLTO più severo per turni ripetuti (aumentato da 5 a 30)
        if turno_str in turni_precedenti:
            # Penalità esponenziale: più vicino è il giorno con lo stesso turno, più forte è la penalità
            count = turni_precedenti.count(turno_str)
            # Calcola la posizione dell'ultimo turno identico
            ultima_posizione = 0
            for i, t in enumerate(reversed(turni_precedenti)):
                if t == turno_str:
                    ultima_posizione = i
                    break
            
            # Penalità più forte se il turno è stato fatto di recente
            punteggio -= 30 * count * (5 - ultima_posizione) / 5
            
            # Se il turno è identico a quello del giorno precedente, penalità extra
            if giorno-1 in turni_assegnati and addetto in turni_assegnati[giorno-1]:
                if f"{turni_assegnati[giorno-1][addetto][0]}-{turni_assegnati[giorno-1][addetto][1]}" == turno_str:
                    punteggio -= 50  # Penalità molto forte per turni identici consecutivi
        
        # Bonus/malus per bilanciare turni mattina/pomeriggio
        turni_mattina = 0
        turni_pomeriggio = 0
        
        for g, turni in turni_assegnati.items():
            if addetto in turni:
                if turni[addetto][0] < "12:00":
                    turni_mattina += 1
                else:
                    turni_pomeriggio += 1
        
        # Incentiva la varietà nei turni (aumento del bonus)
        if turno[0] < "12:00" and turni_mattina < turni_pomeriggio:
            punteggio += 8
        elif turno[0] >= "12:00" and turni_pomeriggio < turni_mattina:
            punteggio += 8
        
        return punteggio

    def _genera_calendario_mensile(self, anno, mese):
        """
        Genera il calendario mensile con queste priorità:
        1. Garantire la copertura COMPLETA dell'orario del negozio (8:00-21:00)
        2. Rispetto assoluto dei vincoli di ore per chi non ha autorizzato straordinari
        3. Ottimizzazione della rotazione dei turni come criterio secondario
        """
        # Ottiene il numero di giorni nel mese
        num_giorni = calendar.monthrange(anno, mese)[1]
        
        # Inizializza il dizionario del calendario
        calendario = {giorno: {} for giorno in range(1, num_giorni + 1)}
        
        # Log per tenere traccia delle ore assegnate a ciascun addetto
        ore_assegnate = {addetto: 0 for addetto in self.addetti}
        
        # Per ogni giorno del mese
        for giorno in range(1, num_giorni + 1):
            data = datetime(anno, mese, giorno)
            data_str = data.strftime('%d-%m')
            
            # Salta i giorni festivi
            if data_str in self.giorni_festivi:
                print(f"Giorno {giorno}: festivo, saltato")
                continue  # Questo continue è correttamente nel ciclo for
            
            print(f"\n=== Pianificazione giorno {giorno} ({data.strftime('%d/%m/%Y')}) ===")
            
            # Lista degli addetti disponibili per il giorno
            addetti_disponibili = []
            addetti_straordinario = []  # Addetti con autorizzazione straordinari
            addetti_no_straordinario = []  # Addetti senza autorizzazione straordinari
            
            for nome, info in self.addetti.items():
                # Controlla ferie
                if data.strftime('%Y-%m-%d') in info['ferie']:
                    print(f"- {nome}: non disponibile (ferie)")
                    continue  # Passa al prossimo addetto
                    
                # Controlla giorni di riposo
                if data.weekday() in info['giorni_riposo']:
                    print(f"- {nome}: non disponibile (riposo settimanale)")
                    continue  # Passa al prossimo addetto
                
                # Calcola ore residue disponibili
                ore_residue = info['ore_max'] - ore_assegnate[nome]
                
                # Per gli addetti senza straordinario, verificare se possono ancora fare turni
                if not info['straordinario']:
                    # Calcola il turno minimo (se non ci sono turni definiti, assumiamo 4 ore)
                    turno_minimo_ore = 4
                    if self.turni_disponibili:
                        ore_min = 24
                        for turno in self.turni_disponibili:
                            inizio = datetime.strptime(turno[0], '%H:%M')
                            fine = datetime.strptime(turno[1], '%H:%M')
                            ore_turno = (fine - inizio).seconds / 3600
                            if ore_turno < ore_min:
                                ore_min = ore_turno
                        turno_minimo_ore = ore_min
                    
                    # Se l'addetto non ha abbastanza ore residue per il turno minimo
                    if ore_residue < turno_minimo_ore:
                        print(f"- {nome}: non disponibile (limite ore raggiunto: {ore_assegnate[nome]:.1f}/{info['ore_max']})")
                        continue  # Passa al prossimo addetto
                    addetti_no_straordinario.append(nome)
                else:
                    addetti_straordinario.append(nome)
                
                # Aggiungi alla lista dei disponibili
                addetti_disponibili.append(nome)
                print(f"- {nome}: disponibile con {ore_residue:.1f} ore residue" + 
                    (" (con straordinario autorizzato)" if info['straordinario'] else ""))
            
            # Se non ci sono addetti disponibili, passa al giorno successivo
            if not addetti_disponibili:
                print(f"AVVISO: Nessun addetto disponibile per il giorno {giorno}!")
                continue  # Questo continue è correttamente nel ciclo for
            
            # Identifica quali fasce orarie devono essere coperte
            # 1. Prepara la copertura oraria (minuti dall'apertura alla chiusura)
            inizio_giornata = datetime.strptime(self.orario_apertura, '%H:%M')
            fine_giornata = datetime.strptime(self.orario_chiusura, '%H:%M')
            
            inizio_min = inizio_giornata.hour * 60 + inizio_giornata.minute
            fine_min = fine_giornata.hour * 60 + fine_giornata.minute
            durata_min = fine_min - inizio_min
            
            # Algoritmo di assegnazione turni
            tentativi = 0
            max_tentativi = 25  # Aumentato per dare più possibilità di trovare una soluzione
            trovata_soluzione_completa = False
            miglior_soluzione = None
            minima_scopertura = durata_min  # Minuti scoperti nella miglior soluzione
            
            while tentativi < max_tentativi and not trovata_soluzione_completa:
                # Inizializza una nuova soluzione
                soluzione_attuale = {}
                copertura = [False] * durata_min  # Minuti coperti
                addetti_disponibili_copia = addetti_disponibili.copy()
                
                # Ore che verrebbero assegnate con questa soluzione
                ore_soluzione = {addetto: 0 for addetto in addetti_disponibili}
                
                # Fase 1: Prima proviamo a coprire l'intera giornata utilizzando tutti gli addetti disponibili
                tentativi_copertura = 0
                while tentativi_copertura < 3:  # Proviamo diversi approcci
                    # Identifica i buchi nella copertura attuale
                    buchi = []
                    inizio_buco = None
                    
                    for i, coperto in enumerate(copertura):
                        if not coperto and inizio_buco is None:
                            inizio_buco = i
                        elif coperto and inizio_buco is not None:
                            buchi.append((inizio_buco, i))
                            inizio_buco = None
                    
                    if inizio_buco is not None:
                        buchi.append((inizio_buco, len(copertura)))
                    
                    # Se non ci sono buchi, abbiamo una copertura completa
                    if not buchi:
                        trovata_soluzione_completa = True
                        break
                    
                    # Per ogni buco, cerchiamo di trovare un turno e un addetto per coprirlo
                    for buco_inizio, buco_fine in buchi:
                        # Converti il buco in orario
                        ora_inizio = inizio_min + buco_inizio
                        ora_fine = inizio_min + buco_fine
                        
                        ora_inizio_str = f"{ora_inizio // 60:02d}:{ora_inizio % 60:02d}"
                        ora_fine_str = f"{ora_fine // 60:02d}:{ora_fine % 60:02d}"
                        
                        print(f"  Trovato buco da coprire: {ora_inizio_str} - {ora_fine_str}")
                        
                        # Cerca turni che potrebbero coprire questo buco (o parte di esso)
                        for turno in self.turni_disponibili:
                            # Converti il turno in minuti
                            t_inizio = datetime.strptime(turno[0], '%H:%M')
                            t_fine = datetime.strptime(turno[1], '%H:%M')
                            
                            t_inizio_min = t_inizio.hour * 60 + t_inizio.minute
                            t_fine_min = t_fine.hour * 60 + t_fine.minute
                            
                            # Verifica se il turno copre almeno parte del buco
                            if (t_inizio_min < ora_fine and t_fine_min > ora_inizio):
                                # Calcola la sovrapposizione
                                inizio_sovrap = max(t_inizio_min, ora_inizio)
                                fine_sovrap = min(t_fine_min, ora_fine)
                                sovrapposizione = fine_sovrap - inizio_sovrap
                                
                                if sovrapposizione <= 0:
                                    continue  # Passa al turno successivo
                                
                                # Cerca un addetto per questo turno
                                miglior_addetto = None
                                miglior_punteggio = -float('inf')
                                
                                # Priorità agli addetti con straordinario autorizzato
                                for addetto in addetti_disponibili_copia:
                                    if addetto in soluzione_attuale:
                                        continue  # Già assegnato, passa al prossimo addetto
                                    
                                    # Calcola ore del turno
                                    ore_turno = (t_fine - t_inizio).seconds / 3600
                                    
                                    # Verifica se l'addetto senza straordinario supererebbe il limite
                                    if (not self.addetti[addetto]['straordinario'] and 
                                        ore_assegnate[addetto] + ore_soluzione[addetto] + ore_turno > self.addetti[addetto]['ore_max']):
                                        continue  # Passa al prossimo addetto
                                    
                                    # Calcola punteggio
                                    punteggio = self._calcola_punteggio_turno(
                                        addetto, turno, giorno, mese, anno, calendario)
                                    
                                    # Bonus per addetti con straordinario autorizzato
                                    if self.addetti[addetto]['straordinario']:
                                        punteggio += 20
                                    
                                    # Bonus maggiore per i turni che coprono più del buco
                                    punteggio += sovrapposizione / 30  # Bonus proporzionale ai minuti coperti
                                    
                                    if punteggio > miglior_punteggio:
                                        miglior_punteggio = punteggio
                                        miglior_addetto = addetto
                                
                                # Se abbiamo trovato un addetto, assegnagli il turno
                                if miglior_addetto:
                                    soluzione_attuale[miglior_addetto] = turno
                                    
                                    # Aggiorna la copertura
                                    t_inizio_rel = max(0, t_inizio_min - inizio_min)
                                    t_fine_rel = min(durata_min, t_fine_min - inizio_min)
                                    
                                    for i in range(t_inizio_rel, t_fine_rel):
                                        copertura[i] = True
                                    
                                    # Aggiorna ore provvisorie
                                    ore_soluzione[miglior_addetto] += ore_turno
                                    
                                    # Rimuovi l'addetto dalla lista dei disponibili se necessario
                                    addetti_disponibili_copia.remove(miglior_addetto)
                                    
                                    # Ricalcola i buchi e riprova
                                    break  # Esce dal ciclo dei turni disponibili
                    
                    # Se abbiamo esaurito gli addetti, interrompi
                    if not addetti_disponibili_copia:
                        break  # Esce dal ciclo dei tentativi di copertura
                    
                    tentativi_copertura += 1
                
                # Calcola quanti minuti sono ancora scoperti
                minuti_scoperti = copertura.count(False)
                
                # Se questa soluzione ha meno minuti scoperti della migliore finora,
                # o è la prima soluzione, memorizzala
                if minuti_scoperti < minima_scopertura or miglior_soluzione is None:
                    minima_scopertura = minuti_scoperti
                    miglior_soluzione = soluzione_attuale.copy()
                    
                    # Se abbiamo trovato una copertura completa, possiamo fermarci
                    if minuti_scoperti == 0:
                        trovata_soluzione_completa = True
                        break  # Esce dal ciclo while principale
                
                # Varia l'approccio per i prossimi tentativi
                tentativi += 1
            
            # Usa la migliore soluzione trovata
            if miglior_soluzione:
                print(f"Soluzione trovata dopo {tentativi+1} tentativi:")
                if minima_scopertura > 0:
                    print(f"ATTENZIONE: {minima_scopertura} minuti rimangono scoperti!")
                
                # Verifica finale che nessun addetto senza straordinario superi il limite
                valida = True
                nuovo_ore_assegnate = ore_assegnate.copy()
                
                for addetto, turno in miglior_soluzione.items():
                    inizio = datetime.strptime(turno[0], '%H:%M')
                    fine = datetime.strptime(turno[1], '%H:%M')
                    ore_turno = (fine - inizio).seconds / 3600
                    nuovo_ore_assegnate[addetto] += ore_turno
                    
                    # Controllo finale rigoroso
                    if (not self.addetti[addetto]['straordinario'] and 
                        nuovo_ore_assegnate[addetto] > self.addetti[addetto]['ore_max']):
                        valida = False
                        print(f"ERRORE: La soluzione farebbe superare il limite a {addetto}")
                        break  # Esce dal ciclo di verifica
                
                if valida:
                    calendario[giorno] = miglior_soluzione
                    
                    # Aggiorna le ore assegnate con i turni effettivi
                    for addetto, turno in miglior_soluzione.items():
                        inizio = datetime.strptime(turno[0], '%H:%M')
                        fine = datetime.strptime(turno[1], '%H:%M')
                        ore_turno = (fine - inizio).seconds / 3600
                        ore_assegnate[addetto] += ore_turno
                        
                        print(f"  Assegnato a {addetto}: {turno[0]}-{turno[1]} ({ore_turno:.1f} ore)")
                else:
                    print("ERRORE: Soluzione non valida, il giorno non verrà coperto!")
            else:
                print("AVVISO: Nessuna soluzione trovata per questo giorno!")
                # NON utilizzare continue qui, siamo già alla fine del ciclo del giorno
        
        # Stampa riepilogo finale
        print("\nRiepilogo ore assegnate:")
        for addetto, ore in ore_assegnate.items():
            info = self.addetti[addetto]
            stato = "OK"
            
            if ore > info['ore_max'] and not info['straordinario']:
                stato = "ERRORE: Superato limite!"
            elif ore > info['ore_max']:
                stato = "Straordinario"
            elif ore < info['ore_contratto']:
                stato = f"Sotto contratto di {info['ore_contratto'] - ore:.1f} ore"
                
            print(f"{addetto}: {ore:.1f} ore / {info['ore_max']} max ({stato})")
        
        return calendario

    def _salva_calendario_excel(self, calendario, anno, mese):
        """Salva il calendario dei turni su file Excel con formattazione migliorata"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Impostazioni di base del foglio
        ws.title = f"Turni {calendar.month_name[mese]} {anno}"
        ws.sheet_view.zoomScale = 85
        
        # Stili comuni
        bordo = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        allineamento = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        # Formattazione header
        header_font = Font(bold=True, size=11, color='000000')
        header_fill = PatternFill(start_color=self.colori['header'], 
                                end_color=self.colori['header'],
                                fill_type='solid')
        
        # Scrivi intestazione
        ws.cell(1, 1, "Giorno").font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 1).border = bordo
        ws.cell(1, 1).alignment = allineamento
        
        # Scrivi nomi addetti nelle colonne
        for col, addetto in enumerate(sorted(self.addetti.keys()), 2):
            cell = ws.cell(1, col, addetto)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = bordo
            cell.alignment = allineamento
        
        # Scrivi i giorni e i turni
        for giorno in range(1, calendar.monthrange(anno, mese)[1] + 1):
            data = datetime(anno, mese, giorno)
            data_str = data.strftime('%d-%m')
            
            # Formattazione riga
            row = giorno + 1
            
            # Scrivi data
            cell_data = ws.cell(row, 1, f"{giorno:02d}/{mese:02d}/{anno}")
            cell_data.border = bordo
            cell_data.alignment = allineamento
            
            # Determina colore di sfondo per il giorno
            if data_str in self.giorni_festivi:
                fill = PatternFill(start_color=self.colori['festivo'], 
                                 end_color=self.colori['festivo'],
                                 fill_type='solid')
            elif data.weekday() >= 5:  # Weekend
                fill = PatternFill(start_color=self.colori['weekend'], 
                                 end_color=self.colori['weekend'],
                                 fill_type='solid')
            else:
                fill = None
            
            if fill:
                cell_data.fill = fill
            
            # Scrivi turni per ogni addetto
            for col, addetto in enumerate(sorted(self.addetti.keys()), 2):
                cell = ws.cell(row, col)
                cell.border = bordo
                cell.alignment = allineamento
                
                # Verifica se è un giorno di ferie
                if data.strftime('%Y-%m-%d') in self.addetti[addetto]['ferie']:
                    cell.value = "FERIE"
                    cell.fill = PatternFill(start_color=self.colori['ferie'], 
                                          end_color=self.colori['ferie'],
                                          fill_type='solid')
                # Verifica se è un giorno di riposo
                elif data.weekday() in self.addetti[addetto]['giorni_riposo']:
                    cell.value = "RIPOSO"
                    cell.fill = PatternFill(start_color=self.colori['riposo'], 
                                          end_color=self.colori['riposo'],
                                          fill_type='solid')
                # Verifica se c'è un turno assegnato
                elif addetto in calendario.get(giorno, {}):
                    turno = calendario[giorno][addetto]
                    cell.value = f"{turno[0]}-{turno[1]}"
                    
                    # Colore diverso per turni mattina/pomeriggio
                    if turno[0] < "12:00":
                        cell.fill = PatternFill(start_color=self.colori['turno_mattina'], 
                                              end_color=self.colori['turno_mattina'],
                                              fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color=self.colori['turno_pomeriggio'], 
                                              end_color=self.colori['turno_pomeriggio'],
                                              fill_type='solid')
                else:
                    cell.value = "-"
        
        # Imposta dimensioni colonne
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Congela la prima riga
        ws.freeze_panes = 'A2'
        
        # Salva il file
        nome_file = f"turni_{anno}_{mese}.xlsx"
        wb.save(nome_file)
        
        # Apri il file
        try:
            os.startfile(nome_file)  # Windows
        except AttributeError:
            import platform
            if platform.system() == "Darwin":  # macOS
                os.system(f"open {nome_file}")
            else:  # Linux
                os.system(f"xdg-open {nome_file}")
        except Exception as e:
            messagebox.showwarning("Attenzione", 
                f"File salvato ma non è stato possibile aprirlo automaticamente.\n{str(e)}")

    def visualizza_statistiche(self):
        """Visualizza le statistiche dei turni"""
        # Verifica che ci siano dati da analizzare
        files_turni = [f for f in os.listdir() if f.startswith('turni_') and f.endswith('.xlsx')]
        if not files_turni:
            messagebox.showinfo("Info", "Nessun dato disponibile per le statistiche")
            return
        
        window = tk.Toplevel(self.root)
        window.title("Statistiche Turni")
        window.geometry("800x600")
        
        # Frame per selezione file
        frame_select = ttk.Frame(window)
        frame_select.pack(pady=10)
        
        ttk.Label(frame_select, text="Seleziona mese:").pack(side=tk.LEFT)
        file_var = tk.StringVar(value=files_turni[0])
        ttk.Combobox(frame_select, textvariable=file_var, 
                    values=files_turni).pack(side=tk.LEFT, padx=5)
        
        # Frame per statistiche
        frame_stats = ttk.Frame(window)
        frame_stats.pack(pady=10, fill=tk.BOTH, expand=True)
        
        def aggiorna_statistiche():
            """Aggiorna le statistiche visualizzate"""
            for widget in frame_stats.winfo_children():
                widget.destroy()
            
            try:
                # Carica dati
                df = pd.read_excel(file_var.get(), index_col=None)
                
                # Calcola statistiche per ogni addetto
                for addetto in df.columns[1:]:  # Salta la colonna Data/Giorno
                    # Conta turni totali
                    turni_totali = 0
                    # Conta ferie e riposi
                    ferie = 0
                    riposi = 0
                    # Ore totali
                    ore_totali = 0
                    # Domeniche lavorate
                    domeniche_lavorate = 0
                    
                    # Analizziamo ogni riga
                    for idx, row in df.iterrows():
                        # Ottieni la data
                        try:
                            data = datetime.strptime(row.iloc[0], '%d/%m/%Y')
                        except (ValueError, TypeError):
                            # Se la colonna non è una data formattata, proviamo a usarla come giorno
                            continue
                        
                        # Ottieni il valore della cella
                        valore_cella = row[addetto]
                        
                        # Verifica il tipo di informazione
                        if isinstance(valore_cella, str):
                            if valore_cella == "FERIE":
                                ferie += 1
                            elif valore_cella == "RIPOSO":
                                riposi += 1
                            elif "-" in valore_cella:  # È un turno (formato "HH:MM-HH:MM")
                                turni_totali += 1
                                
                                # Calcola ore turno
                                try:
                                    inizio, fine = valore_cella.split("-")
                                    inizio = inizio.strip()
                                    fine = fine.strip()
                                    
                                    # Verifica che il formato sia corretto
                                    if ":" in inizio and ":" in fine:
                                        ore_inizio = datetime.strptime(inizio, '%H:%M')
                                        ore_fine = datetime.strptime(fine, '%H:%M')
                                        ore_turno = (ore_fine - ore_inizio).seconds / 3600
                                        ore_totali += ore_turno
                                        
                                        # Verifica se è una domenica
                                        if data.weekday() == 6:  # 6 = domenica
                                            domeniche_lavorate += 1
                                except Exception as e:
                                    print(f"Errore nell'analisi del turno {valore_cella}: {e}")
                    
                    # Crea frame per addetto
                    frame_addetto = ttk.LabelFrame(frame_stats, text=addetto)
                    frame_addetto.pack(fill=tk.X, padx=5, pady=5)
                    
                    # Inizio statistiche
                    ttk.Label(frame_addetto, 
                            text=f"Turni totali nel mese: {turni_totali}").pack(padx=5, pady=2)
                    ttk.Label(frame_addetto, 
                            text=f"Giorni di ferie: {ferie}").pack(padx=5, pady=2)
                    ttk.Label(frame_addetto, 
                            text=f"Giorni di riposo: {riposi}").pack(padx=5, pady=2)
                    ttk.Label(frame_addetto, 
                            text=f"Domeniche lavorate: {domeniche_lavorate}").pack(padx=5, pady=2)
                    ttk.Label(frame_addetto, 
                            text=f"Ore totali lavorate: {ore_totali:.1f}").pack(padx=5, pady=2)
                    
                    # Verifica rispetto monte ore
                    if addetto in self.addetti:
                        ore_contratto = self.addetti[addetto]['ore_contratto']
                        ore_max = self.addetti[addetto]['ore_max']
                        
                        if ore_totali < ore_contratto:
                            ttk.Label(frame_addetto, 
                                    text=f"⚠️ Ore sotto contratto di {ore_contratto - ore_totali:.1f} ore",
                                    foreground='orange').pack(padx=5, pady=2)
                        elif ore_totali > ore_max and not self.addetti[addetto]['straordinario']:
                            ttk.Label(frame_addetto, 
                                    text=f"⚠️ Superato limite ore di {ore_totali - ore_max:.1f} ore",
                                    foreground='red').pack(padx=5, pady=2)
            
            except Exception as e:
                messagebox.showerror("Errore", f"Errore nell'analisi dei dati: {str(e)}")
                print(f"Dettaglio errore: {e}")
        
        # Bottone per aggiornare statistiche
        ttk.Button(frame_select, text="Aggiorna Statistiche", 
                  command=aggiorna_statistiche).pack(side=tk.LEFT, padx=5)
        
        # Aggiorna statistiche iniziali
        aggiorna_statistiche()

    def genera_pianificazione(self):
        """Genera la pianificazione dei turni per il mese selezionato"""
        if not self.addetti or not self.turni_disponibili:
            messagebox.showerror("Errore", "Inserire prima addetti e turni disponibili")
            return
        
        window = tk.Toplevel(self.root)
        window.title("Genera Pianificazione")
        window.geometry("400x300")
        
        # Frame per selezione periodo
        frame_periodo = ttk.LabelFrame(window, text="Seleziona Periodo")
        frame_periodo.pack(pady=10, padx=10, fill='x')
        
        # Anno
        ttk.Label(frame_periodo, text="Anno:").grid(row=0, column=0, padx=5)
        anno_var = tk.StringVar(value=str(datetime.now().year))
        ttk.Entry(frame_periodo, textvariable=anno_var, width=6).grid(row=0, column=1)
        
        # Mese
        ttk.Label(frame_periodo, text="Mese:").grid(row=0, column=2, padx=5)
        mesi = list(calendar.month_name)[1:]  # Esclude il primo elemento vuoto
        mese_var = tk.StringVar(value=mesi[datetime.now().month - 1])
        ttk.Combobox(frame_periodo, textvariable=mese_var, 
                    values=mesi, width=10).grid(row=0, column=3)
        
        def genera():
            """Genera i turni per il mese selezionato"""
            try:
                anno = int(anno_var.get())
                mese = mesi.index(mese_var.get()) + 1
                
                # Genera il calendario mensile
                calendario = self._genera_calendario_mensile(anno, mese)
                
                # Salva su Excel
                self._salva_calendario_excel(calendario, anno, mese)
                
                messagebox.showinfo("Successo", "Pianificazione generata e salvata")
                window.destroy()
                
            except ValueError:
                messagebox.showerror("Errore", "Data non valida")
            except Exception as e:
                messagebox.showerror("Errore", f"Si è verificato un errore: {str(e)}")
        
        ttk.Button(window, text="Genera Pianificazione", 
                  command=genera).pack(pady=20)
    
    def run(self):
        """Avvia l'applicazione"""
        self.root.mainloop()

# Avvio dell'applicazione
if __name__ == "__main__":
    app = GestioneTurni()
    app.run()